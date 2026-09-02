"""Junta saida_1..5.json com o dossiê todos.json e grava um Excel: uma aba por classificação."""
import collections, datetime as dt, json, os, sys, unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))

CANON = {
    "QUEIMADO": "QUEIMADO", "AVARIADO": "AVARIADO", "FURTO": "FURTO",
    "ABALROAMENTO": "ABALROAMENTO", "PREVENTIVO/PROGRAMADO": "PREVENTIVO/PROGRAMADO",
    "MELHORIA DE POSTO": "MELHORIA DE POSTO", "REMANEJAMENTO": "REMANEJAMENTO",
    "DIVISAO DE CIRCUITO": "DIVISÃO DE CIRCUITO",
    "TAPE/REGULARIZACAO DE TENSAO": "TAPE/REGULARIZAÇÃO DE TENSÃO",
    "PARTICULAR": "PARTICULAR", "AUXILIAR DE RELIGADOR": "AUXILIAR DE RELIGADOR",
    "POSTE/REDE (NAO E O TRAFO)": "POSTE/REDE (não é o trafo)",
    "SEM TROCA (NAO SUBSTITUIDO)": "SEM TROCA (não substituído)",
    "SS CANCELADA/DUPLICADA": "SS CANCELADA/DUPLICADA", "INCONCLUSIVO": "INCONCLUSIVO",
}
ORDEM = ["QUEIMADO", "AVARIADO", "FURTO", "ABALROAMENTO", "PREVENTIVO/PROGRAMADO",
         "MELHORIA DE POSTO", "REMANEJAMENTO", "DIVISÃO DE CIRCUITO",
         "TAPE/REGULARIZAÇÃO DE TENSÃO", "PARTICULAR", "AUXILIAR DE RELIGADOR",
         "POSTE/REDE (não é o trafo)", "SEM TROCA (não substituído)",
         "SS CANCELADA/DUPLICADA", "INCONCLUSIVO"]
NOME_ABA = {"TAPE/REGULARIZAÇÃO DE TENSÃO": "Tape e tensão",
            "POSTE/REDE (não é o trafo)": "Poste ou rede",
            "SEM TROCA (não substituído)": "Sem troca",
            "SS CANCELADA/DUPLICADA": "SS cancelada ou duplicada",
            "PREVENTIVO/PROGRAMADO": "Preventivo ou programado",
            "DIVISÃO DE CIRCUITO": "Divisão de circuito",
            "MELHORIA DE POSTO": "Melhoria de posto",
            "AUXILIAR DE RELIGADOR": "Auxiliar de religador"}


def canon(c):
    s = unicodedata.normalize("NFKD", str(c or "")).encode("ascii", "ignore").decode().upper().strip()
    return CANON.get(s, str(c or "INCONCLUSIVO"))


def main():
    casos = {c["ss"]: c for c in json.load(open(os.path.join(BASE, "todos.json"), encoding="utf-8"))}
    vereditos, faltando = {}, []
    for i in range(1, 6):
        p = os.path.join(BASE, f"saida_{i}.json")
        if not os.path.exists(p):
            faltando.append(i); continue
        try:
            d = json.load(open(p, encoding="utf-8"))
        except Exception as e:
            faltando.append(i); print(f"saida_{i}.json ilegível: {e}"); continue
        for v in d:
            vereditos[v["ss"]] = v
    sem = [s for s in casos if s not in vereditos]
    print(f"lidos {len(vereditos)}/205 · lotes faltando {faltando} · SS sem veredito {len(sem)}")
    if faltando or sem:
        print("AINDA INCOMPLETO — não gravei o Excel")
        return 1

    linhas = []
    for ss, c in casos.items():
        v = vereditos[ss]
        cr = c.get("critica") or {}
        st = c.get("site") or {}
        campo = c.get("campo") or {}
        linhas.append({
            "SS": ss, "OS": c.get("os"), "Obra": c.get("obra"),
            "Transformador": c.get("trafo"), "Abertura": c.get("abertura"),
            "Término": c.get("termino"),
            "Classificação (leitura caso a caso)": canon(v.get("classificacao")),
            "Decisão": v.get("decisao"), "Confiança": v.get("confianca"),
            "Justificativa": v.get("justificativa"), "Evidência literal": v.get("evidencia"),
            "Diverge do site": "SIM" if v.get("diverge_do_site") else "não",
            "Classificação do dono": c.get("classificacao_do_dono"),
            "Tipo SS": c.get("tipo_ss"), "Origem SS": c.get("origem_ss"),
            "Defeito SS": c.get("defeito_ss"), "Situação": c.get("situacao"),
            "Localidade": c.get("localidade"), "Alimentador": c.get("alimentador"),
            "Equipe": c.get("equipe"), "kVA retirado": c.get("pot_ret"), "kVA instalado": c.get("pot_inst"),
            "Crítica · resultado": cr.get("resultado"), "Crítica · ocorrência": cr.get("ocorrencia"),
            "Crítica · início": cr.get("inicio"), "Crítica · fim": cr.get("fim"),
            "Crítica · papel": cr.get("papel"), "Crítica · distância (h)": cr.get("dist_h"),
            "Crítica · causa": cr.get("causa"), "Crítica · subcausa": cr.get("subcausa"),
            "Crítica · clientes": cr.get("clientes"), "Crítica · observação": cr.get("obs"),
            "Site · cascata": st.get("cascata"), "Site · gatilho": st.get("gatilho"),
            "Site · confirmado": st.get("confirmado"), "Site · motivo": st.get("motivo"),
            "Série retirada": campo.get("serie_retirado"), "Série instalada": campo.get("serie_instalado"),
            "Tombamento retirado": campo.get("tombamento_retirado"),
            "Tombamento instalado": campo.get("tombamento_instalado"),
            "Vazamento de óleo": campo.get("vazamento_oleo"),
            "Provável motivo (campo)": campo.get("provavel_motivo"),
            "Texto da SS": c.get("texto_ss"), "Texto da OS": c.get("texto_os"),
        })
    linhas.sort(key=lambda l: str(l["Abertura"]))
    ordem = list(linhas[0].keys())

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Resumo"
    ws.append(["As 205 SS marcadas ANÁLISE DE EXPURGOS — leitura caso a caso da SS e da OS"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"Gerado em {dt.datetime.now():%d/%m/%Y %H:%M} · fonte Trafo_v1_3.xlsx (aba BASE SS_OS, coluna A) + Crítica + site"])
    ws.append([])
    ws.append(["Classificação", "SS", "EXPURGAR", "MANTER", "confiança baixa", "diverge do site"])
    for c in ws[4]: c.font = Font(bold=True)
    por = collections.defaultdict(list)
    for l in linhas: por[l["Classificação (leitura caso a caso)"]].append(l)
    for k in ORDEM:
        g = por.get(k)
        if not g: continue
        ws.append([k, len(g),
                   sum(1 for l in g if l["Decisão"] == "EXPURGAR"),
                   sum(1 for l in g if l["Decisão"] == "MANTER"),
                   sum(1 for l in g if str(l["Confiança"]).lower() == "baixa"),
                   sum(1 for l in g if l["Diverge do site"] == "SIM")])
    ws.append(["Total", len(linhas),
               sum(1 for l in linhas if l["Decisão"] == "EXPURGAR"),
               sum(1 for l in linhas if l["Decisão"] == "MANTER"),
               sum(1 for l in linhas if str(l["Confiança"]).lower() == "baixa"),
               sum(1 for l in linhas if l["Diverge do site"] == "SIM")])
    for c in ws[ws.max_row]: c.font = Font(bold=True)
    ws.append([])

    ws.append(["Contra a classificação do dono (aba Expurgos 309)"]); ws[ws.max_row][0].font = Font(bold=True)
    ws.append(["Classificação do dono", "SS", "mesma leitura", "leitura diferente"])
    for c in ws[ws.max_row]: c.font = Font(bold=True)
    pd = collections.defaultdict(list)
    for l in linhas:
        if l["Classificação do dono"]: pd[l["Classificação do dono"]].append(l)
    for k in sorted(pd):
        g = pd[k]
        ig = sum(1 for l in g if canon(k) == l["Classificação (leitura caso a caso)"])
        ws.append([k, len(g), ig, len(g) - ig])
    ws.append([f"(sem classificação do dono)", sum(1 for l in linhas if not l["Classificação do dono"])])
    ws.append([])

    ws.append(["Contra o que o site decidiu"]); ws[ws.max_row][0].font = Font(bold=True)
    ws.append(["Site · cascata", "SS", "eu digo EXPURGAR", "eu digo MANTER"])
    for c in ws[ws.max_row]: c.font = Font(bold=True)
    ps = collections.defaultdict(list)
    for l in linhas: ps[l["Site · cascata"] or "(não está no site)"].append(l)
    for k in sorted(ps, key=lambda x: -len(ps[x])):
        g = ps[k]
        ws.append([k, len(g), sum(1 for l in g if l["Decisão"] == "EXPURGAR"), sum(1 for l in g if l["Decisão"] == "MANTER")])
    ws.append([])
    for t in ["Leitura ao lado do caso: nada aqui recalcula o 1.305 nem o 1.582 — o site continua como está.",
              "Cada SS foi lida com o texto da SS, o texto da OS, os campos do formulário de campo (série e tombamento retirado/instalado, vazamento de óleo, provável motivo), o resultado na Crítica e a decisão do site.",
              "EXPURGAR = a leitura do caso diz que a troca não foi por queima/avaria do transformador. MANTER = a troca é queima ou avaria e o caso fica no indicador.",
              "Confiança baixa = o texto não fecha; precisa de foto, laudo ou confirmação da equipe antes de decidir.",
              "Diverge do site = minha leitura do caso não bate com a cascata do site; a coluna Justificativa diz por quê."]:
        ws.append([t])
    ws.column_dimensions["A"].width = 46
    for c in "BCDEF": ws.column_dimensions[c].width = 20

    def aba(nome, dados):
        w = wb.create_sheet(nome[:31])
        w.append(ordem)
        for c in w[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F3864")
            c.alignment = Alignment(wrap_text=True, vertical="center")
        for d in dados: w.append([d.get(c) for c in ordem])
        for i, c in enumerate(ordem, 1):
            w.column_dimensions[get_column_letter(i)].width = min(60, max(12, len(c) + 2))
        w.freeze_panes = "B2"
        if dados: w.auto_filter.ref = w.dimensions

    aba("Todas as 205", linhas)
    for k in ORDEM:
        if por.get(k): aba(NOME_ABA.get(k, k.capitalize()), por[k])
    div = [l for l in linhas if l["Diverge do site"] == "SIM"]
    if div: aba("Divergem do site", div)
    bx = [l for l in linhas if str(l["Confiança"]).lower() == "baixa"]
    if bx: aba("Confiança baixa", bx)

    saida = os.path.join(BASE, f"Analise_205_SS_ANALISE_DE_EXPURGOS.xlsx")
    wb.save(saida)
    print(saida)
    for k in ORDEM:
        if por.get(k): print(f"  {k}: {len(por[k])}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
