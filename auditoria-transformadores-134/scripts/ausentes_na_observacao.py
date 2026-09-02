"""Os ausentes da Crítica, procurados dentro do texto de OBSERVACAO.

Pedido do dono: "verificar se podem estar na observação da Crítica". Quem não aparece nas
três colunas de ativo (nem pelo trafo nem pela chave gêmea) pode ter sido escrito à mão pelo
executante no campo de observação — "trafo 5700121151 queimado", "TR-5700121151", ou só os
oito dígitos finais com outro prefixo.

O que se procura, por SS ausente:
  · o código inteiro do transformador (10 dígitos);
  · os 8 dígitos finais, com qualquer prefixo (a numeração muda entre SS, KML e Crítica);
  · o número da SS (por exemplo "00003/2026" ou "3/2026" junto da sigla da equipe).
Cada achado diz em que ocorrência está, se a abertura da SS cai na janela dela (−1h/+24h
sobre a ocorrência inteira, como no site), qual elemento consta como problema e o trecho.

Entradas: a planilha de saída de analise_ss_critica_chave.py (aba Ausentes e Fora da janela)
          e a Crítica (zip jan–jun + extras).
Saída:    JSON e planilha com um achado por linha, mais o resumo por SS.
"""

import argparse
import collections
import datetime as dt
import json
import os
import re
import sys
import zipfile

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TRAFO_PREF = ("57", "53", "52", "42", "51", "54", "56", "55")
AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.join(AQUI, "..")
PUB = os.path.join(RAIZ, "public")
sys.path.insert(0, AQUI)
from analise_ss_critica_chave import ler_critica_arquivo, data, txt, ANTES, DEPOIS  # noqa: E402


def carregar_passos(zip_path, extras):
    passos = []
    with zipfile.ZipFile(zip_path) as z:
        for info in sorted(z.infolist(), key=lambda i: i.filename):
            passos += ler_critica_arquivo(info.filename, z.read(info).decode("latin-1"))
    for e in extras:
        with open(e, encoding="latin-1") as fh:
            passos += ler_critica_arquivo(os.path.basename(e), fh.read())
    ocs = {}
    for p in passos:
        o = ocs.setdefault(p["oc"], {"ini": p["ini"], "fim": p["fim"], "passos": 0})
        o["passos"] += 1
        if p["ini"] and (o["ini"] is None or p["ini"] < o["ini"]):
            o["ini"] = p["ini"]
        if p["fim"] and (o["fim"] is None or p["fim"] > o["fim"]):
            o["fim"] = p["fim"]
    return passos, ocs


def ler_casos(planilha, abas):
    wb = openpyxl.load_workbook(planilha, read_only=True)
    casos = []
    for aba in abas:
        ws = wb[aba]
        rows = ws.iter_rows(values_only=True)
        hdr = next(rows)
        if hdr[0] == "nenhuma":
            continue
        H = {h: i for i, h in enumerate(hdr)}
        for r in rows:
            casos.append({
                "grupo": aba, "ss": txt(r[H["SS"]]), "os": txt(r[H["OS"]]), "trafo": txt(r[H["Transformador"]]),
                "abertura": r[H["Abertura da SS"]], "localidade": txt(r[H["Localidade"]]),
                "alimentador": txt(r[H["Alimentador"]]), "origem": txt(r[H["Origem SS"]]),
                "resultado": txt(r[H["Resultado"]]), "site": txt(r[H["Site · cascata"]]),
                "site_gatilho": txt(r[H["Site · gatilho de exclusão"]]),
            })
    return casos


def padroes_da_ss(numero_ss):
    """'DOLP-RD-PA 00003/2026' → os textos que identificam ESTA SS: a sigla da equipe (com ou
    sem hífens e espaços) seguida do número, com ou sem zeros à esquerda. O número sozinho
    ("3/2026") não vale: a primeira rodada casou "ENC-RD-PS26/2026" com a SS 26 de outra
    equipe e "DOLP-RD-PA 295/2026" com a ETO-RD-GR 295 — SS diferentes, mesmo número."""
    m = re.match(r"\s*([A-Z]+)-([A-Z]+)-([A-Z]+)\s+(\d{3,6})/(\d{4})", numero_ss)
    if not m:
        return []
    a, b, c, n, ano = m.groups()
    return [(a, b, c, n, ano)]


def regex_da_ss(numero_ss):
    ps = padroes_da_ss(numero_ss)
    if not ps:
        return None
    a, b, c, n, ano = ps[0]
    return re.compile(rf"{a}[\s\-]*{b}[\s\-]*{c}[\s\-:]*0*{int(n)}\s*/\s*{ano}(?!\d)", re.I)


def principal():
    ap = argparse.ArgumentParser()
    ap.add_argument("--planilha", default=os.path.join(RAIZ, "..", "SS_x_Critica_trafo_e_chave_2026.xlsx"))
    ap.add_argument("--abas", nargs="*", default=["Ausentes", "Fora da janela"])
    ap.add_argument("--critica-extra", nargs="*", default=[])
    ap.add_argument("--saida", default=os.path.join(RAIZ, "..", "Ausentes_na_Observacao_da_Critica.xlsx"))
    a = ap.parse_args()

    casos = ler_casos(a.planilha, a.abas)
    print(f"casos: {len(casos)} · " + " · ".join(f"{k}={v}" for k, v in collections.Counter(c['grupo'] for c in casos).items()))
    passos, ocs = carregar_passos(os.path.join(PUB, "bases", "originais", "Original_Critica_Interrupcoes.zip"), a.critica_extra)
    com_obs = [p for p in passos if p["obs"]]
    print(f"passos: {len(passos)} · com observação: {len(com_obs)}")

    # índice: para cada caso, os alvos de texto
    alvos = []  # (regex, caso, tipo)
    for c in casos:
        cod = c["trafo"]
        f8 = cod[2:]
        alvos.append((re.compile(r"(?<!\d)" + re.escape(cod) + r"(?!\d)"), c, "código inteiro"))
        # os 8 finais com prefixo de transformador (ou sem prefixo) falam do trafo; com prefixo
        # 79/03/02/67/33/88 falam do religador ou da chave do mesmo ponto — outro equipamento.
        alvos.append((re.compile(r"(?<!\d)(\d{0,2})" + re.escape(f8) + r"(?!\d)"), c, "8 dígitos finais"))
        rx_ss = regex_da_ss(c["ss"])
        if rx_ss:
            alvos.append((rx_ss, c, "número da SS"))

    # varrer só os passos com observação; para não passar 300 regex em 76 mil linhas à toa,
    # pré-filtra pelos 8 dígitos finais com busca de string
    # o mesmo transformador pode estar em duas SS (troca repetida): cada final aponta para
    # TODAS as SS dele — a primeira versão guardava só a última e perdeu duas menções
    finais = collections.defaultdict(list)
    for c in casos:
        finais[c["trafo"][2:]].append(c)
    achados = []
    for p in com_obs:
        obs = p["obs"]
        candidatos = [c for f8, cs in finais.items() if f8 in obs for c in cs]
        # números de SS: pré-filtro barato pelo "número/ano" e confirmação pela sigla no regex
        for c in casos:
            ps = padroes_da_ss(c["ss"])
            if ps and c not in candidatos:
                n, ano = ps[0][3], ps[0][4]
                if f"{int(n)}/{ano}" in obs or f"{n}/{ano}" in obs:
                    candidatos.append(c)
        for c in candidatos:
            tipos = []
            for rx, cc, tipo in alvos:
                if cc is not c:
                    continue
                if tipo == "8 dígitos finais":
                    prefixos = {m8.group(1) for m8 in rx.finditer(obs)}
                    if not prefixos:
                        continue
                    if any(pf in TRAFO_PREF or pf == "" or len(pf) < 2 for pf in prefixos):
                        tipos.append("8 dígitos finais (trafo)")
                    outros = sorted(pf for pf in prefixos if len(pf) == 2 and pf not in TRAFO_PREF)
                    if outros:
                        tipos.append("outro equipamento de mesmo final (" + "/".join(outros) + ")")
                elif rx.search(obs):
                    tipos.append(tipo)
            if not tipos:
                continue
            if "código inteiro" in tipos and "8 dígitos finais (trafo)" in tipos:
                tipos.remove("8 dígitos finais (trafo)")
            fala_do_trafo = any(t in ("código inteiro", "8 dígitos finais (trafo)", "número da SS") for t in tipos)
            o = ocs[p["oc"]]
            ab = c["abertura"]
            na_janela = bool(o["ini"] and o["fim"] and ab and (o["ini"] - ANTES) <= ab <= (o["fim"] + DEPOIS))
            if o["ini"] and o["fim"] and ab and not na_janela:
                dist = (o["ini"] - ab).total_seconds() / 3600 if ab < o["ini"] else (ab - o["fim"]).total_seconds() / 3600
            else:
                dist = 0.0 if na_janela else None
            m = re.search(re.escape(c["trafo"][2:]), obs)
            i = m.start() if m else 0
            trecho = obs[max(0, i - 120): i + 160]
            achados.append({
                "Grupo": c["grupo"], "SS": c["ss"], "Transformador": c["trafo"], "Abertura da SS": ab,
                "Como apareceu": " + ".join(tipos), "Fala do trafo": "SIM" if fala_do_trafo else "NÃO — outro equipamento", "Ocorrência": p["oc"], "Arquivo": p["arquivo"],
                "Início da ocorrência": o["ini"], "Fim da ocorrência": o["fim"], "Passos": o["passos"],
                "Na janela": "SIM" if na_janela else "NÃO", "Distância à janela (h)": None if dist is None else round(dist, 1),
                "Elemento com problema": p["problema"], "Elemento interrompido": p["interrompido"], "Elemento fechado": p["fechado"],
                "Causa": p["causa"], "Subcausa": p["subcausa"], "Clientes": p["clientes"],
                "Trecho da observação": trecho, "Observação inteira": obs[:1000],
                "Site · cascata": c["site"], "Site · gatilho": c["site_gatilho"], "Resultado anterior": c["resultado"],
            })

    # resumo por SS
    por_ss = collections.defaultdict(list)
    for x in achados:
        por_ss[x["SS"]].append(x)
    resumo = []
    for c in casos:
        todos = por_ss.get(c["ss"], [])
        xs = [x for x in todos if x["Fala do trafo"] == "SIM"]
        so_outro = [x for x in todos if x["Fala do trafo"] != "SIM"]
        na = [x for x in xs if x["Na janela"] == "SIM"]
        resumo.append({
            "Grupo": c["grupo"], "SS": c["ss"], "Transformador": c["trafo"], "Abertura da SS": c["abertura"],
            "Localidade": c["localidade"], "Origem SS": c["origem"], "Resultado anterior": c["resultado"],
            "Menções na observação": len(xs), "…na janela": len(na),
            "Veredito": ("MENCIONADO NA JANELA" if na else ("MENCIONADO FORA DA JANELA" if xs else
                         ("SÓ OUTRO EQUIPAMENTO DE MESMO FINAL" if so_outro else "NÃO MENCIONADO"))),
            "Menções a outro equipamento de mesmo final": len(so_outro),
            "Ocorrências (na janela)": "; ".join(sorted({x["Ocorrência"] for x in na})),
            "Elemento com problema (na janela)": "; ".join(sorted({x["Elemento com problema"] for x in na if x["Elemento com problema"]})),
            "Primeiro trecho": (na or xs)[0]["Trecho da observação"] if xs else "",
            "Site · cascata": c["site"], "Site · gatilho": c["site_gatilho"],
        })

    cont = collections.Counter((r["Grupo"], r["Veredito"]) for r in resumo)
    json.dump({"casos": len(casos), "achados": achados, "resumo": resumo}, open(a.saida.replace(".xlsx", ".json"), "w", encoding="utf-8"),
              ensure_ascii=False, default=str, indent=1)

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Resumo"
    ws.append(["Ausentes e fora da janela, procurados no texto de OBSERVACAO da Crítica"]); ws["A1"].font = Font(bold=True)
    ws.append([f"Gerado em {dt.datetime.now():%d/%m/%Y %H:%M} · {len(casos)} SS · {len(com_obs)} passos com observação varridos"])
    ws.append([])
    ws.append(["Grupo", "Veredito", "SS"]); [setattr(c, "font", Font(bold=True)) for c in ws[4]]
    for (g, v), n in sorted(cont.items()):
        ws.append([g, v, n])
    ws.append([])
    for t in ["Busca: código inteiro do trafo; 8 dígitos finais com qualquer prefixo; número da SS (nnnnn/2026).",
              "Na janela = abertura da SS entre (início − 1h) e (fim + 24h) da ocorrência inteira, como no site.",
              "Mencionado não é o mesmo que ter defeito: o elemento com problema da ocorrência está na coluna ao lado, e o trecho mostra o que o executante escreveu.",
              "Leitura ao lado do caso; não mexe no 1.305 nem no 1.582."]:
        ws.append([t])
    ws.column_dimensions["A"].width = 40; ws.column_dimensions["B"].width = 32

    def aba(nome, dados):
        w = wb.create_sheet(nome)
        if not dados:
            w.append(["nenhuma"]); return
        cols = list(dados[0].keys()); w.append(cols)
        for c in w[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F3864"); c.alignment = Alignment(wrap_text=True)
        for d in dados:
            w.append([d[c] for c in cols])
        for i, c in enumerate(cols, 1):
            w.column_dimensions[get_column_letter(i)].width = min(60, max(12, len(c) + 2))
        for row in w.iter_rows(min_row=2):
            for c in row:
                if isinstance(c.value, dt.datetime):
                    c.number_format = "dd/mm/yyyy hh:mm"
        w.freeze_panes = "B2"; w.auto_filter.ref = w.dimensions

    aba("Por SS", resumo)
    aba("Achados", achados)
    aba("Achados na janela", [x for x in achados if x["Na janela"] == "SIM" and x["Fala do trafo"] == "SIM"])
    aba("Outro equipamento", [x for x in achados if x["Fala do trafo"] != "SIM"])
    wb.save(a.saida)
    print(a.saida)
    for (g, v), n in sorted(cont.items()):
        print(f"  {g} · {v}: {n}")


if __name__ == "__main__":
    principal()
