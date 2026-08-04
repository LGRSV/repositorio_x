#!/usr/bin/env python3
"""Gera Bases_Gerais.xlsx — as bases da auditoria num arquivo só, uma aba por base.

Pedido dele: "preciso de um outro excel com as bases gerais para fins de pesquisa organizada por
aba". Os arquivos já existiam separados — seis downloads, seis janelas do Excel, e nenhuma forma
de cruzar dois deles sem abrir os dois. Aqui é um arquivo: cada base vira uma aba, com filtro
automático ligado e a primeira linha congelada, para procurar e cruzar sem sair de dentro.

Não recalcula nada e não resume nada: é cópia fiel do que já foi gerado. Se uma base mudar, este
script é rodado de novo e a cópia acompanha — a aba Índice carimba a data e o tamanho de cada uma,
então uma cópia velha se denuncia sozinha.

Uso:  python3 gerar_planilha_bases.py
"""
import datetime
import json
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASES = os.path.join(RAIZ, "public", "bases")
FLUXO = os.path.join(RAIZ, "public", "fluxo-1510.json")
SAIDA = os.path.join(BASES, "Bases_Gerais.xlsx")

TINTA = PatternFill("solid", fgColor="1f2a37")

# arquivo de origem, nome da aba, e para que serve — a ordem é a da esteira, não a alfabética:
# a SS abre o caso, a interrupção prova o fato, o atendimento corrobora, a obra e o material
# comprovam a troca. Quem for pesquisar segue essa ordem sem precisar de mapa.
FONTES = [
    ("Base_SS_OS.xlsx", "1 · SS e OS",
     "As 1.510 solicitações com todos os campos do cadastro e o texto integral da SS e da OS. "
     "É a chave de tudo: a coluna SS liga esta aba a todas as outras."),
    ("Base_Interrupcoes.xlsx", "2 · Interrupções",
     "A linha da Crítica de cada SS que casou com uma ocorrência: datas, duração, clientes "
     "interrompidos, elemento do defeito, causa, subcausa e observação de campo."),
    ("Base_Atendimentos_TMAE.xlsx", "3 · Atendimentos",
     "A linha do TMAE: cronologia completa, os quatro tempos, equipe e a observação de quem "
     "executou. Marcador, não prova — a chave do TMAE é o elemento onde o defeito foi aberto."),
    ("Base_Obras_SIGCO.xlsx", "4 · Obras e SIGCO",
     "O cadastro da obra de cada caso: classe, natureza, tipo, projeto SIGCO, empreiteira, setor, "
     "valores previsto e realizado, e as datas de cada fase."),
    ("Base_Material.xlsx", "5 · Material da obra",
     "Item a item do que saiu do almoxarifado, com código, descrição, quantidade prevista e "
     "realizada e valor. É aqui que se lê se um transformador foi movimentado."),
    ("Base_Esteira_Completa.xlsx", "6 · Esteira completa",
     "Uma linha por SS com a posição na esteira, o motivo, a decisão, a causa confirmada, o "
     "gatilho da exclusão com a frase que o explica e o intervalo inteiro da ocorrência."),
    ("Material_Pendente.xlsx", "7 · Material pendente",
     "Os casos que o export de material não responde, com a obra de cada um. Só a aba do caso a "
     "caso entra aqui; o arquivo separado traz o resumo e a lista de obras a extrair."),
]


def copia(origem, wb, titulo, aba_origem=None):
    """Copia uma aba inteira, valores e larguras. Sem fórmula: base de pesquisa é dado, não conta."""
    wo = openpyxl.load_workbook(origem, read_only=True, data_only=True)
    ws_o = wo[aba_origem] if aba_origem else wo[wo.sheetnames[0]]
    ws = wb.create_sheet(titulo[:31])
    for linha in ws_o.iter_rows(values_only=True):
        ws.append(list(linha))
    largos = {}
    for i, linha in enumerate(ws.iter_rows(values_only=True)):
        if i > 400:  # amostra suficiente para largura; ler 15 mil linhas para isso é desperdício
            break
        for j, v in enumerate(linha, start=1):
            largos[j] = max(largos.get(j, 10), len(str(v)) if v is not None else 0)
    for j, w in largos.items():
        ws.column_dimensions[get_column_letter(j)].width = min(58, max(11, w + 2))
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = TINTA
    ws.freeze_panes = "A2"
    if ws.max_row > 1 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    linhas, colunas = ws.max_row - 1, ws.max_column
    wo.close()
    return linhas, colunas


def main():
    with open(FLUXO, encoding="utf-8") as fh:
        carimbo = json.load(fh)["meta"].get("revisado_em") or str(datetime.date.today())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    indice = wb.create_sheet("Índice")

    resumo = []
    for arquivo, titulo, nota in FONTES:
        caminho = os.path.join(BASES, arquivo)
        if not os.path.exists(caminho):
            print(f"  ! falta {arquivo} — pulado")
            continue
        aba_origem = "Caso a caso" if arquivo == "Material_Pendente.xlsx" else None
        linhas, colunas = copia(caminho, wb, titulo, aba_origem)
        resumo.append([titulo, linhas, colunas, arquivo, nota])
        print(f"  {titulo:26} {linhas:>6} linhas × {colunas} colunas")

    indice.append([f"Bases da auditoria de transformadores — cópia fiel gerada em {carimbo}"])
    indice.merge_cells("A1:E1")
    c = indice.cell(row=1, column=1)
    c.font = Font(bold=True, size=12)
    indice.append(["Nenhum valor foi recalculado ou resumido: cada aba é a base como ela é, com "
                   "filtro automático ligado. A coluna SS liga todas elas."])
    indice.merge_cells("A2:E2")
    indice.cell(row=2, column=1).font = Font(italic=True, color="44506a")
    indice.append([])
    indice.append(["Aba", "Linhas", "Colunas", "Arquivo de origem", "Para que serve"])
    for c in indice[4]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = TINTA
    for l in resumo:
        indice.append(l)
    for j, w in enumerate([26, 9, 9, 30, 96], start=1):
        indice.column_dimensions[get_column_letter(j)].width = w
    for linha in indice.iter_rows(min_row=5):
        linha[4].alignment = Alignment(wrap_text=True, vertical="top")
        indice.row_dimensions[linha[0].row].height = 34
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(indice)))

    wb.save(SAIDA)
    print(f"\n{SAIDA}")
    print(f"  {len(resumo)} abas · {sum(l[1] for l in resumo)} linhas no total")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
