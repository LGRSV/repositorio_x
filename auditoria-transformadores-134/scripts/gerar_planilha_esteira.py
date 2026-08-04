#!/usr/bin/env python3
"""Regera Base_Esteira_Completa.xlsx a partir do fluxo-1510.json.

A planilha existia sem gerador: era montada à mão a cada rodada, e envelhecia calada. O
invariante 19 nasceu justamente de a planilha dizer SAÍDA=874 enquanto o dado já dizia 884 —
quem baixasse o arquivo levava outra história para a reunião. Agora ela é derivada, e uma
rodada da esteira seguida desta é o bastante para as duas contarem o mesmo.

Uma coluna nova: "Fora da esteira" e "Por que foi excluída", porque a exclusão deixou de ser
etapa da cascata e passou a ser porta anterior a ela.
"""
import datetime
import json
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FLUXO = os.path.join(RAIZ, "public", "fluxo-1510.json")
SAIDA = os.path.join(RAIZ, "public", "bases", "Base_Esteira_Completa.xlsx")

# (título, chave no registro). A ordem é a de leitura: identificação, veredito, prova, origem.
COLUNAS = [
    ("SS", "ss"), ("OS", "os"), ("Obra", "obra"), ("Ativo", "trafo"),
    ("Abertura", "abertura"), ("Localidade", "localidade"),
    ("Cascata", "cascata"), ("Motivo", "cascata_motivo"),
    ("Fora da esteira", "fora_da_esteira"), ("Gatilho da exclusão", "expurgo_gatilho"),
    ("Por que foi excluída", "exclusao_porque"),
    ("Decisão", "decisao"), ("Confirmado", "confirmado"),
    ("Fato", "fato"), ("Leitura", "leitura"),
    ("Categoria pelo texto", "categoria_texto"), ("Categoria gravada", "categoria_gravada"),
    ("Sob suspeita", "sob_suspeita"), ("Sinais de atenção", "suspeitas"),
    ("Ocorrência", "oc_num"), ("Ocorrência início", "oc_ini"), ("Ocorrência fim", "oc_fim"),
    ("Duração da ocorrência (h)", "oc_dur_h"), ("Passos da ocorrência", "oc_passos"),
    ("Atendimento", "at_num"), ("Atendimento início", "at_ini"),
    ("Equipe do TMAE", "at_equipe"), ("Observação do executante", "at_obs"),
    ("Deslocamento", "deslocamento"), ("Equipe", "equipe_ss"),
    ("Trafos no material", "trafos_material"), ("Ressalvas", "ressalvas"),
    ("Alertas de obra", "e4_alertas"), ("Solicitante", "solicitante"), ("Origem", "origem"),
    ("Revisado em (Brasília)", "revisado_em"),
]


def main():
    with open(FLUXO, encoding="utf-8") as fh:
        fluxo = json.load(fh)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Esteira"
    ws.append([t for t, _ in COLUNAS])
    cab = PatternFill("solid", fgColor="1f2a37")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = cab
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    for r in fluxo["registros"]:
        ws.append([r.get(k) if r.get(k) is not None else "" for _, k in COLUNAS])
    # largura por coluna: o suficiente para ler sem abrir célula por célula
    for i, (titulo, chave) in enumerate(COLUNAS, start=1):
        largo = chave in ("cascata_motivo", "exclusao_porque", "at_obs", "suspeitas",
                          "ressalvas", "e4_alertas")
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 60 if largo else max(
            12, min(28, len(titulo) + 4))
    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    tam = os.path.getsize(SAIDA)
    print(f"{SAIDA} · {len(fluxo['registros'])} linhas · {len(COLUNAS)} colunas · "
          f"{tam / 1_000_000:.2f} MB")
    print(f"carimbo do dado: {fluxo['meta'].get('revisado_em') or datetime.date.today()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
