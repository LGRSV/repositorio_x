#!/usr/bin/env python3
"""Gera Sem_Interrupcao_Critica.xlsx — os 137 que a Crítica não sustenta, caso a caso.

Pedido dele: "traga o restante das faixas de tempo num excel, quero acompanhar esses sem
interrupção pela crítica". A planilha existe para acompanhamento — é a lista que ele vai abrir
para conferir um a um, então cada linha traz o que a base CRUA diz, não o rótulo da tela.

Três abas, e a divisão entre elas é a única que a base sustenta:

  · FORA DA JANELA — tem defeito aberto no próprio transformador, em outra data. Cada linha traz
    a distância em horas e em dias, e a faixa de tempo.
  · AUSENTES — o código não aparece na Crítica em papel nenhum, nos sete meses. Não há ocorrência
    para mostrar; a única ligação é o teste do vizinho, que é hipótese e vem marcado como tal.
  · APARECE SEM DEFEITO NELE — o grupo que ele mandou deixar de lado por ora: aparece na Crítica
    como interrompido ou manobrado, nunca como o elemento com defeito. Fica em aba própria para
    não contaminar a distribuição de tempo das outras, porque para estes não existe distância.

Os três papéis de cada código — COM DEFEITO, INTERROMPIDO e MANOBRADO — são relidos aqui dos
arquivos originais com csv.QUOTE_NONE. Nenhum campo já calculado decide o que entra em cada aba.

Uso:  python3 gerar_planilha_sem_interrupcao.py [pasta com os .txt da Crítica]
"""
import collections
import csv
import datetime
import glob
import json
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FLUXO = os.path.join(RAIZ, "public", "fluxo-1510.json")
SAIDA = os.path.join(RAIZ, "public", "bases", "Sem_Interrupcao_Critica.xlsx")
TINTA = PatternFill("solid", fgColor="1f2a37")

FAIXAS = [
    ("até 24 horas", lambda x: x <= 24),
    ("1 a 7 dias", lambda x: 24 < x <= 168),
    ("7 a 30 dias", lambda x: 168 < x <= 720),
    ("1 a 3 meses", lambda x: 720 < x <= 2160),
    ("mais de 3 meses", lambda x: x > 2160),
]


def t(v):
    return "" if v is None else str(v).strip()


def faixa(h):
    for nome, teste in FAIXAS:
        if teste(h):
            return nome
    return ""


def le_critica(padroes):
    """Devolve, por código de ativo, em que papéis ele aparece na Crítica."""
    defeito = collections.defaultdict(list)
    interrompido = collections.defaultdict(set)
    manobrado = collections.defaultdict(set)
    arquivos = [a for p in padroes for a in glob.glob(p)]
    for f in arquivos:
        with open(f, encoding="latin-1", newline="") as fh:
            rd = csv.reader(fh, delimiter=";", quoting=csv.QUOTE_NONE)
            cab = next(rd)
            i = {c: k for k, c in enumerate(cab)}
            for row in rd:
                if len(row) < len(cab):
                    continue
                oc = row[i["NUM_SEQ_OPER_INIC_HDE"]].strip()
                if row[i["COD_ELE_REDE_PROBLEMA"]].strip() == "TR":
                    cod = row[i["COD_ELE_PROBLEMA"]].strip()
                    if cod:
                        defeito[cod].append((row[i["DTA_ABERT"]].strip(), oc,
                                             row[i["DES_CAUSA_INTER_CAU"]].strip()))
                if row[i["COD_ELE_INTERROMPIDO"]].strip():
                    interrompido[row[i["COD_ELE_INTERROMPIDO"]].strip()].add(oc)
                if row[i["COD_ELE_FECHADO"]].strip():
                    manobrado[row[i["COD_ELE_FECHADO"]].strip()].add(oc)
    return defeito, interrompido, manobrado, len(arquivos)


def aba(wb, titulo, colunas, linhas, nota, larguras):
    ws = wb.create_sheet(titulo[:31])
    ws.append([nota])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(colunas))
    c = ws.cell(row=1, column=1)
    c.font = Font(italic=True, color="44506a")
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 46
    ws.append([])
    ws.append(colunas)
    for c in ws[3]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = TINTA
    for l in linhas:
        ws.append(l)
    for j, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A4"
    if ws.max_row > 3:
        ws.auto_filter.ref = f"A3:{get_column_letter(len(colunas))}{ws.max_row}"
    return ws


def main(argv):
    padroes = argv or [
        "/tmp/claude-0/-home-user/74dc9c64-5026-54ee-a81e-173d2f38a735/scratchpad/crit/*.txt",
        "/root/.claude/uploads/74dc9c64-5026-54ee-a81e-173d2f38a735/*Critica*.txt",
    ]
    defeito, interrompido, manobrado, n_arq = le_critica(padroes)
    if not defeito:
        print("não achei os arquivos da Crítica — passe o caminho como argumento")
        return 2
    with open(FLUXO, encoding="utf-8") as fh:
        fluxo = json.load(fh)
    R = fluxo["registros"]
    carimbo = fluxo["meta"].get("revisado_em") or str(datetime.date.today())
    PAROU = {"fora_da_janela", "sem_interrupcao"}
    presos = [r for r in R if t(r.get("cascata")) == "EXCLUÍDA" and t(r.get("expurgo_gatilho")) in PAROU]

    com_def, ausentes, sem_def = [], [], []
    for r in presos:
        cod = t(r.get("trafo"))
        if defeito.get(cod):
            com_def.append(r)
        elif interrompido.get(cod) or manobrado.get(cod):
            sem_def.append(r)
        else:
            ausentes.append(r)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ------------------------------------------------------------------ resumo
    dist = sorted(abs(float(r.get("oc_dist_h") or 0)) for r in com_def)
    linhas = [
        ["Presos pela Crítica", len(presos), "não têm interrupção que sustente o caso"],
        ["", "", ""],
        ["Fora da janela — tem defeito nele, em outra data", len(com_def), "é para estes que existe distância"],
        ["Ausentes — não aparecem em papel nenhum", len(ausentes), "não há ocorrência para mostrar"],
        ["Aparecem, mas nunca com defeito nele", len(sem_def), "interrompido ou manobrado — sem distância"],
        ["", "", ""],
        ["FAIXAS DE TEMPO — só os que têm registro em outra data", "", ""],
    ]
    for nome, teste in FAIXAS:
        n = sum(1 for x in dist if teste(x))
        linhas.append([f"  {nome}", n, f"{round(100 * n / max(1, len(com_def)))}% dos {len(com_def)}"])
    if dist:
        linhas += [
            ["", "", ""],
            ["  a mais perto da janela", round(dist[0], 1), "horas"],
            ["  mediana", round(dist[len(dist) // 2] / 24, 1), "dias"],
            ["  a mais longe", round(dist[-1] / 24), "dias"],
        ]
    aba(wb, "Resumo", ["O que é", "Quantas", "Leitura"], linhas,
        f"Os {len(presos)} que a Crítica não sustenta — conferidos contra os {n_arq} arquivos "
        f"originais da base, relidos com csv.QUOTE_NONE. Gerado de fluxo-1510.json em {carimbo}. "
        "A divisão em três é a que a base sustenta, não a que o rótulo diz.",
        [52, 11, 40])

    # ------------------------------------------------------- fora da janela
    COLS = ["SS", "Transformador", "Localidade", "Alimentador", "Abertura da SS", "Faixa de tempo",
            "Distância (horas)", "Distância (dias)", "Ocorrência mais próxima", "Data dela",
            "Causa dela", "Ocorrências com defeito nele", "Texto da SS", "Obra", "Trafos no material"]
    linhas = []
    for r in sorted(com_def, key=lambda x: abs(float(x.get("oc_dist_h") or 0))):
        cod = t(r.get("trafo"))
        h = abs(float(r.get("oc_dist_h") or 0))
        regs = defeito.get(cod, [])
        linhas.append([t(r.get("ss")), cod, t(r.get("localidade")), t(r.get("alimentador")),
                       t(r.get("abertura")), faixa(h), round(h, 1), round(h / 24, 1),
                       t(r.get("oc_num")), t(r.get("oc_ini")), t(r.get("oc_causa")),
                       len(regs), t(r.get("desc_ss"))[:300], t(r.get("obra_descricao")),
                       r.get("trafos_material")])
    aba(wb, "Fora da janela", COLS, linhas,
        f"{len(com_def)} solicitações em que a Crítica registra defeito aberto NO PRÓPRIO "
        "transformador, em data que não cabe na janela desta SS. A distância é medida da abertura "
        "da SS até a borda mais próxima do intervalo inteiro da ocorrência — a régua da peneira. "
        "Ordenadas da mais perto para a mais longe.",
        [22, 15, 22, 14, 18, 16, 15, 15, 20, 18, 26, 13, 60, 30, 12])

    # ------------------------------------------------------------- ausentes
    COLS2 = ["SS", "Transformador", "Localidade", "Alimentador", "Abertura da SS",
             "Ocorrência no dossiê", "Atendimento do TMAE", "Como está vinculado",
             "Teste do vizinho", "Texto da SS", "Obra", "Trafos no material"]
    linhas = []
    for r in sorted(ausentes, key=lambda x: t(x.get("ss"))):
        viz = t(r.get("vizinho"))
        tem_viz = bool(viz) and not viz.startswith("Nada")
        linhas.append([t(r.get("ss")), t(r.get("trafo")), t(r.get("localidade")),
                       t(r.get("alimentador")), t(r.get("abertura")),
                       t(r.get("oc_num")) or "— nenhuma", t(r.get("at_num")) or "— nenhum",
                       "hipótese do vizinho" if tem_viz else "nada — sobe para campo",
                       viz, t(r.get("desc_ss"))[:300], t(r.get("obra_descricao")),
                       r.get("trafos_material")])
    aba(wb, "Ausentes da Crítica", COLS2, linhas,
        f"{len(ausentes)} solicitações cujo código não aparece na Crítica em papel nenhum — nem "
        "com defeito, nem interrompido, nem manobrado — nos sete meses do acervo. As colunas de "
        "ocorrência e de atendimento vêm vazias de propósito: não existe registro, e mostrar “a "
        "ocorrência mais próxima” seria inventar vínculo. A única ligação é o teste do vizinho, "
        "que aponta ocorrência em OUTRO ativo do mesmo alimentador — hipótese, não prova.",
        [22, 15, 22, 14, 18, 20, 20, 24, 66, 60, 30, 12])

    # ------------------------------------------- aparece, mas sem defeito nele
    COLS3 = ["SS", "Transformador", "Localidade", "Abertura da SS", "Papéis na Crítica",
             "Texto da SS", "Obra", "Trafos no material"]
    linhas = []
    for r in sorted(sem_def, key=lambda x: t(x.get("ss"))):
        cod = t(r.get("trafo"))
        pap = []
        if interrompido.get(cod): pap.append(f"interrompido em {len(interrompido[cod])}")
        if manobrado.get(cod): pap.append(f"manobrado em {len(manobrado[cod])}")
        linhas.append([t(r.get("ss")), cod, t(r.get("localidade")), t(r.get("abertura")),
                       " · ".join(pap), t(r.get("desc_ss"))[:300], t(r.get("obra_descricao")),
                       r.get("trafos_material")])
    aba(wb, "Aparece sem defeito nele", COLS3, linhas,
        f"{len(sem_def)} solicitações em que o código APARECE na Crítica, mas nunca como o "
        "elemento com defeito: só como interrompido ou manobrado. Para estas não existe distância "
        "até a janela, e por isso elas não entram na distribuição de tempo da primeira aba. Hoje "
        "o site as classifica junto das de fora da janela — é o assunto que ficou em aberto.",
        [22, 15, 22, 18, 34, 60, 30, 12])

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    print(SAIDA)
    print(f"  {len(presos)} presos · {len(com_def)} fora da janela · {len(ausentes)} ausentes · "
          f"{len(sem_def)} aparecem sem defeito nele")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
