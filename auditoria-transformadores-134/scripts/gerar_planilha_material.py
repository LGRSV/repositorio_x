#!/usr/bin/env python3
"""Gera Material_Pendente.xlsx — os casos sem conferência de material, com a obra de cada um.

Pedido dele: "preciso daqueles casos que precisamos ver o que não tem verificação de material,
preciso de uma planilha com a obra desses casos". A planilha existe para virar pedido: quem for
puxar o SIAGO precisa da lista de OBRAS, não da lista de SS — uma obra pode carregar mais de uma
solicitação, e pedir por SS faria a extração vir repetida.

São dois problemas diferentes debaixo do mesmo rótulo, e a planilha não os mistura:

  · A OBRA EXISTE e não está no export de material que temos. Aqui não falta prova: falta a
    extração. É a aba "Obras a extrair", e é a fila mais barata de fechar desta auditoria.
  · A OBRA NUNCA FOI GERADA. Não há o que extrair. Estes precisam de outra providência —
    abrir a obra, ou aceitar que a prova de troca não virá.

Regra da casa: nenhum número é digitado. Tudo sai do fluxo-1510.json na hora.

Uso:  python3 gerar_planilha_material.py
"""
import datetime
import json
import os
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FLUXO = os.path.join(RAIZ, "public", "fluxo-1510.json")
SAIDA = os.path.join(RAIZ, "public", "bases", "Material_Pendente.xlsx")

TINTA = PatternFill("solid", fgColor="1f2a37")
BORDA = Border(bottom=Side(style="thin", color="d4dae2"))


def t(v):
    return "" if v is None else str(v).strip()


def aba(wb, titulo, colunas, linhas, nota=""):
    """Uma aba com cabeçalho escuro, largura ajustada e a nota de rodapé."""
    ws = wb.create_sheet(titulo[:31])
    if nota:
        ws.append([nota])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(colunas)))
        c = ws.cell(row=1, column=1)
        c.font = Font(italic=True, color="44506a")
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[1].height = 42
        ws.append([])
    ini = ws.max_row + 1
    ws.append(colunas)
    for c in ws[ini]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = TINTA
    for l in linhas:
        ws.append(l)
    for i, cab in enumerate(colunas, start=1):
        largo = max([len(str(cab))] + [len(str(l[i - 1])) for l in linhas if i <= len(l)] or [10])
        ws.column_dimensions[get_column_letter(i)].width = min(60, max(12, largo + 3))
    ws.freeze_panes = ws.cell(row=ini + 1, column=1)
    return ws


def main():
    with open(FLUXO, encoding="utf-8") as fh:
        fluxo = json.load(fh)
    R = fluxo["registros"]
    carimbo = fluxo["meta"].get("revisado_em") or str(datetime.date.today())

    pend = [r for r in R if t(r.get("material_conferido")) != "SIM"]
    com_obra = [r for r in pend if t(r.get("obra"))]
    sem_obra = [r for r in pend if not t(r.get("obra"))]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ------------------------------------------------------------------ resumo
    por_destino = Counter(t(r.get("cascata")) for r in pend)
    linhas = [
        ["Solicitações sem conferência de material", len(pend), "o export de material não responde por elas"],
        ["  · a obra existe e está fora do export", len(com_obra), "falta a extração do SIAGO, não a prova"],
        ["  · a obra nunca foi gerada", len(sem_obra), "não há o que extrair — precisa de outra providência"],
        ["", "", ""],
        ["Obras distintas a extrair", len({t(r.get("obra")) for r in com_obra}), "é por aqui que o pedido deve ser feito"],
        ["", "", ""],
    ]
    for destino, n in por_destino.most_common():
        linhas.append([f"Onde estão hoje · {destino}", n, ""])
    linhas += [
        ["", "", ""],
        ["Marcadas como pendentes só do SIAGO", sum(1 for r in pend if t(r.get("pendente_siago")) == "SIM"),
         "obra com número, descrição e enquadramento — só falta a extração"],
        ["Retidas na esteira por falta desta prova", sum(1 for r in pend if t(r.get("cascata")).startswith("RETIDO")),
         "são as únicas que a conferência do material ainda pode mover"],
    ]
    aba(wb, "Resumo", ["O que é", "Quantas", "Leitura"], linhas,
        f"Casos sem conferência de material — gerado de fluxo-1510.json em {carimbo}. "
        "São dois problemas diferentes: obra que existe e está fora do export (falta a extração) "
        "e obra que nunca foi gerada (não há o que extrair).")

    # ------------------------------------------------- obras a extrair, uma por linha
    # A extração se pede por OBRA. Agrupar aqui evita que a mesma obra seja pedida duas vezes
    # quando ela carrega mais de uma solicitação.
    grupos = defaultdict(list)
    for r in com_obra:
        grupos[t(r.get("obra"))].append(r)
    linhas = []
    for obra, rs in sorted(grupos.items(), key=lambda kv: -len(kv[1])):
        p = rs[0]
        linhas.append([
            obra, len(rs), " · ".join(t(x.get("ss")) for x in rs),
            t(p.get("obra_descricao")), t(p.get("obra_status")), t(p.get("obra_tipo")),
            t(p.get("obra_setor")), t(p.get("obra_empreiteira")),
            float(p.get("obra_realizado") or 0), t(p.get("sigco")), t(p.get("sigco_descricao")),
            " · ".join(t(x.get("cascata")) for x in rs),
            "SIM" if any(t(x.get("pendente_siago")) == "SIM" for x in rs) else "",
        ])
    aba(wb, "Obras a extrair", [
        "Obra", "SS na obra", "Solicitações", "Descrição da obra", "Status", "Tipo",
        "Setor", "Empreiteira", "Valor realizado", "SIGCO", "Descrição SIGCO",
        "Onde a SS está hoje", "Só falta o SIAGO",
    ], linhas,
        f"{len(linhas)} obras carregando {len(com_obra)} solicitações. Todas existem no cadastro "
        "de obras e nenhuma aparece no export de material que temos. É esta a lista do pedido de "
        "extração — pedir por SS traria a mesma obra repetida.")

    # ------------------------------------------------------------- caso a caso
    COLS = [
        ("SS", "ss"), ("OS", "os"), ("Transformador", "trafo"), ("Localidade", "localidade"),
        ("Alimentador", "alimentador"), ("Abertura", "abertura"), ("Idade da SS (dias)", "idade_ss"),
        ("Situação", "situacao"), ("Onde está hoje", "cascata"), ("Motivo do material", "e3_motivo"),
        ("Obra", "obra"), ("Descrição da obra", "obra_descricao"), ("Status da obra", "obra_status"),
        ("Setor", "obra_setor"), ("Empreiteira", "obra_empreiteira"),
        ("Valor realizado", "obra_realizado"), ("SIGCO", "sigco"),
        ("Trafos no material", "trafos_material"), ("Só falta o SIAGO", "pendente_siago"),
        ("Categoria de exclusão", "expurgo_gatilho"), ("O que o texto diz", "categoria_texto"),
        ("Descrição da SS", "desc_ss"),
    ]
    linhas = [[t(r.get(c)) for _, c in COLS] for r in
              sorted(pend, key=lambda r: (t(r.get("e3_motivo")), t(r.get("obra")), t(r.get("ss"))))]
    aba(wb, "Caso a caso", [n for n, _ in COLS], linhas,
        f"As {len(pend)} solicitações sem conferência de material, uma por linha, com a obra de "
        "cada uma quando existe. Ordenadas pelo motivo e depois pela obra.")

    # ------------------------------------------------------- as que não têm obra
    linhas = [[t(r.get(c)) for _, c in COLS if c not in ("obra_descricao", "obra_status",
                                                          "obra_setor", "obra_empreiteira",
                                                          "obra_realizado", "sigco")]
              for r in sorted(sem_obra, key=lambda r: t(r.get("ss")))]
    aba(wb, "Sem obra gerada", [n for n, c in COLS if c not in ("obra_descricao", "obra_status",
                                                                "obra_setor", "obra_empreiteira",
                                                                "obra_realizado", "sigco")], linhas,
        f"{len(sem_obra)} solicitações em que a obra nunca foi gerada. Para estas não adianta pedir "
        "extração: não existe obra no cadastro. A providência é outra — abrir a obra, ou aceitar "
        "que a prova de troca não virá e decidir o caso por outro caminho.")

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    print(f"{SAIDA}")
    print(f"  {len(pend)} sem conferência de material · {len(com_obra)} com obra "
          f"({len(grupos)} obras distintas) · {len(sem_obra)} sem obra gerada")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
