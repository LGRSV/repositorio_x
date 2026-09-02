"""SS de transformador de 2026 × Crítica, pelo trafo e — só se não achar — pela chave gêmea.

Pedido do dono (01/09): "Pega a Base de SS OS Ano 2026, código operativo 42, 52, 53, 57, com
data de abertura de janeiro a julho. Pesquise os ativos na Crítica da mesma forma que é feita
no site. Adicione somente a procura pela chave do ativo (começa com 03 e os 8 últimos dígitos
são iguais) caso não encontre pelo número operativo do transformador nas 3 colunas."

O MÉTODO É O DO SITE, sem inventar nada:
  · as três colunas de ativo da Crítica: COD_ELE_PROBLEMA, COD_ELE_INTERROMPIDO, COD_ELE_FECHADO;
  · a ocorrência é o conjunto de passos com o mesmo NUM_SEQ_OPER_INIC_HDE, e a janela é medida
    contra o intervalo inteiro dela — do primeiro passo aberto ao último fechado;
  · casa quando a abertura da SS cai entre (início − 1 h) e (fim + 24 h);
  · a chave gêmea só é procurada para quem não aparece pelo trafo em papel nenhum.

O QUE ESTE SCRIPT NÃO É: não é a esteira. Não lê TMAE, material, texto. É uma leitura ao lado
do caso, comparada no fim com o que o site decidiu para a mesma SS.

Entradas:
  public/bases/originais/Original_SS_TRAFOS_V4.xlsx (aba BASE GERAL) ou, se existir, a base
    @ do ano (BASE_SS_OS*.txt) — passe pelo argumento --ss
  Crítica: public/bases/originais/Original_Critica_Interrupcoes.zip (jan–jun) + qualquer
    Critica*.txt extra (julho) em --critica-extra
Saída: planilha .xlsx no caminho de --saida
"""

import argparse
import collections
import datetime as dt
import os
import re
import sys
import zipfile

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.join(AQUI, "..")
PUB = os.path.join(RAIZ, "public")
sys.path.insert(0, os.path.join(AQUI, "analise-mensal"))
import base  # noqa: E402  — o mesmo leitor da base @ que o site usa para os meses

PREFIXOS = ("42", "52", "53", "57")
PAPEIS = ("COD_ELE_PROBLEMA", "COD_ELE_INTERROMPIDO", "COD_ELE_FECHADO")
ANTES = dt.timedelta(hours=1)
DEPOIS = dt.timedelta(hours=24)
INICIO, FIM = dt.datetime(2026, 1, 1), dt.datetime(2026, 8, 1)

txt = lambda v: "" if v is None else str(v).strip()


def data(v):
    if isinstance(v, dt.datetime):
        return v
    s = txt(v)
    for f in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(s[:19] if "T" not in s else s.replace("T", " ")[:19], f)
        except ValueError:
            continue
    return None


# ── SS ──────────────────────────────────────────────────────────────────────
def ler_ss_v4(caminho):
    wb = openpyxl.load_workbook(caminho, read_only=True)
    ws = wb["BASE GERAL"]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    H = {h: i for i, h in enumerate(hdr) if h}
    g = lambda r, k: txt(r[H[k]]) if k in H else ""
    saida = []
    for r in rows:
        saida.append({
            "ss": g(r, "NUMERO_SS"), "os": g(r, "NUMERO_OS"), "obra": g(r, "NUM_OBRA"),
            "trafo": g(r, "NUM_TRAFO"), "abertura": data(r[H["DATA_ABERTURA_SS"]]),
            "origem": g(r, "ORIGEM_SS"), "defeito": g(r, "DEFEITO_SS"), "tipo_ss": g(r, "TIPOSS"),
            "situacao": g(r, "SITUACAO_SS"), "criticidade": g(r, "CRITICIDADE_SS"),
            "localidade": g(r, "LOCALIDADE"), "alimentador": g(r, "ALIMENTADOR"),
            "equipe": g(r, "COD_EQUIPE"), "pot_ret": g(r, "POTENCIA_RET"), "pot_inst": g(r, "POTENCIA_INST"),
            "descricao": g(r, "DESCRIPTION")[:400],
        })
    return "Original_SS_TRAFOS_V4.xlsx · BASE GERAL", saida


def ler_ss_arroba(*arquivos):
    cab, regs = base.ler_ss_os(*arquivos)
    g = base.indexar(cab)
    saida = []
    for r in regs:
        saida.append({
            "ss": g(r, "NUMERO_SS"), "os": g(r, "NUMERO_OS"), "obra": g(r, "NUM_OBRA"),
            "trafo": g(r, "NUM_TRAFO"), "abertura": data(g(r, "DATA_ABERTURA_SS")),
            "origem": g(r, "ORIGEM_SS"), "defeito": g(r, "DEFEITO_SS"), "tipo_ss": g(r, "TIPOSS"),
            "situacao": g(r, "SITUACAO_SS"), "criticidade": g(r, "CRITICIDADE_SS"),
            "localidade": g(r, "LOCALIDADE"), "alimentador": g(r, "ALIMENTADOR"),
            "equipe": g(r, "COD_EQUIPE"), "pot_ret": g(r, "POTENCIA_RET"), "pot_inst": g(r, "POTENCIA_INST"),
            "descricao": (g(r, "DESCRIPTION_SS") or g(r, "DESCRIPTION"))[:400],
            "descricao_os": g(r, "DESCRICAO_OS")[:400],
        })
    return " + ".join(os.path.basename(a) for a in arquivos), saida


# ── Crítica ────────────────────────────────────────────────────────────────
def ler_critica_arquivo(nome, bruto):
    """Devolve lista de dicts por passo. Latin-1, ';', cabeçalho pelo nome (janeiro troca a
    ordem de duas colunas de índice; por nome isso não importa)."""
    linhas = bruto.split("\n")
    cab = [c.strip() for c in linhas[0].split(";")]
    I = {c: n for n, c in enumerate(cab)}
    for c in PAPEIS + ("NUM_SEQ_OPER_INIC_HDE", "DTA_ABERT", "DTA_FECH"):
        if c not in I:
            raise RuntimeError(f"{nome}: falta a coluna {c}")
    passos = []
    for l in linhas[1:]:
        if not l.strip():
            continue
        c = l.split(";")
        if len(c) < len(cab):
            continue
        passos.append({
            "arquivo": nome, "oc": txt(c[I["NUM_SEQ_OPER_INIC_HDE"]]),
            "ini": data(c[I["DTA_ABERT"]]), "fim": data(c[I["DTA_FECH"]]),
            "problema": txt(c[I["COD_ELE_PROBLEMA"]]), "interrompido": txt(c[I["COD_ELE_INTERROMPIDO"]]),
            "fechado": txt(c[I["COD_ELE_FECHADO"]]),
            "causa": txt(c[I.get("DES_CAUSA_INTER_CAU", 0)]) if "DES_CAUSA_INTER_CAU" in I else "",
            "subcausa": txt(c[I["DES_SUB_CAUSA_INTER_SCR"]]) if "DES_SUB_CAUSA_INTER_SCR" in I else "",
            "clientes": txt(c[I["QTD_CONS_INTER_FAT"]]) if "QTD_CONS_INTER_FAT" in I else "",
            "duracao": txt(c[I["DURACAO"]]) if "DURACAO" in I else "",
            "obs": txt(c[I["OBSERVACAO"]]) if "OBSERVACAO" in I else "",
        })
    return passos


def carregar_critica(zip_path, extras):
    passos = []
    with zipfile.ZipFile(zip_path) as z:
        for info in sorted(z.infolist(), key=lambda i: i.filename):
            bruto = z.read(info).decode("latin-1")
            p = ler_critica_arquivo(info.filename, bruto)
            print(f"  {info.filename}: {len(p)} passos")
            passos += p
    for e in extras:
        with open(e, encoding="latin-1") as fh:
            bruto = fh.read()
        p = ler_critica_arquivo(os.path.basename(e), bruto)
        print(f"  {os.path.basename(e)}: {len(p)} passos")
        passos += p
    # ocorrência = mesmo NUM_SEQ_OPER_INIC_HDE; a janela é do primeiro passo ao último
    ocs = {}
    por_codigo = collections.defaultdict(lambda: collections.defaultdict(set))  # codigo → oc → papéis
    for p in passos:
        o = ocs.setdefault(p["oc"], {"oc": p["oc"], "ini": p["ini"], "fim": p["fim"], "passos": 0,
                                     "causa": p["causa"], "subcausa": p["subcausa"],
                                     "clientes": p["clientes"], "duracao": p["duracao"],
                                     "obs": p["obs"], "arquivo": p["arquivo"]})
        o["passos"] += 1
        if p["ini"] and (o["ini"] is None or p["ini"] < o["ini"]):
            o["ini"] = p["ini"]
        if p["fim"] and (o["fim"] is None or p["fim"] > o["fim"]):
            o["fim"] = p["fim"]
        if not o["obs"] and p["obs"]:
            o["obs"] = p["obs"]
        for papel in ("problema", "interrompido", "fechado"):
            if p[papel]:
                por_codigo[p[papel]][p["oc"]].add(papel)
    return ocs, por_codigo


ORDEM_PAPEL = {"problema": 0, "interrompido": 1, "fechado": 2}


def papeis_texto(papeis):
    """'problema+interrompido+fechado', sempre nesta ordem — julho vinha em outra e virava
    balde separado em qualquer pivô por papel."""
    return "+".join(sorted(set(papeis), key=lambda p: ORDEM_PAPEL.get(p, 9)))


def distancia_a_janela(abertura, ini, fim):
    """Horas que a abertura da SS ficou FORA da janela [ini − 1 h, fim + 24 h]; 0 dentro."""
    if abertura < ini - ANTES:
        return (ini - ANTES - abertura).total_seconds() / 3600
    if abertura > fim + DEPOIS:
        return (abertura - fim - DEPOIS).total_seconds() / 3600
    return 0.0


def procurar(codigo, abertura, ocs, por_codigo):
    """Devolve (status, melhor_ocorrencia, papeis, delta_h, n_ocorrencias).

    Entre as ocorrências que casam, a melhor é a em que o trafo é o elemento com PROBLEMA e,
    empatando, a de início mais perto da abertura. Fora da janela, a de menor distância à
    borda. Ocorrência sem DTA_FECH (ainda aberta na extração) usa o início como fim, em vez
    de ser descartada — descartar podia mandar a busca para a chave gêmea sem motivo."""
    achadas = por_codigo.get(codigo)
    if not achadas or not abertura:
        return "AUSENTE", None, "", None, len(achadas or {})
    dentro, fora = [], []
    for oc_id, papeis in achadas.items():
        o = ocs[oc_id]
        ini = o["ini"] or o["fim"]
        fim = o["fim"] or o["ini"]
        if not ini:
            continue
        d = distancia_a_janela(abertura, ini, fim)
        chave_ordem = (0 if "problema" in papeis else 1, abs((abertura - ini).total_seconds()))
        (dentro if d == 0 else fora).append((chave_ordem if d == 0 else (d,), o, papeis_texto(papeis)))
    if dentro:
        _, o, pap = min(dentro, key=lambda t: t[0])
        return "SIM", o, pap, None, len(achadas)
    if fora:
        (d,), o, pap = min(fora, key=lambda t: t[0])
        return "fora da janela", o, pap, round(d, 1), len(achadas)
    return "AUSENTE", None, "", None, len(achadas)


# ── o site, para comparar ──────────────────────────────────────────────────
def carregar_site():
    import json
    cam = os.path.join(PUB, "fluxo-1582.json")
    if not os.path.exists(cam):
        return {}
    f = json.load(open(cam, encoding="utf-8"))
    return {r["ss"]: r for r in f["registros"]}


def principal():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ss", nargs="*", default=None, help="base @ do ano (BASE_SS_OS*.txt); sem ela usa a V4")
    ap.add_argument("--critica-extra", nargs="*", default=[], help="Critica*.txt fora do zip (julho)")
    ap.add_argument("--saida", default=os.path.join(RAIZ, "..", "SS_x_Critica_trafo_e_chave_2026.xlsx"))
    ap.add_argument("--julho-json", default=os.path.join(PUB, "julho-2026.json"),
                    help="extração por SS da Crítica de julho guardada no site; '' para não usar")
    a = ap.parse_args()

    print("SS:")
    if a.ss:
        fonte_ss, ss = ler_ss_arroba(*a.ss)
    else:
        fonte_ss, ss = ler_ss_v4(os.path.join(PUB, "bases", "originais", "Original_SS_TRAFOS_V4.xlsx"))
    print(f"  {fonte_ss}: {len(ss)} linhas")
    # A Crítica bruta de julho (Critica__072026.txt) foi enviada em sessão anterior e se perdeu
    # nos reinícios do contêiner. O que sobrou dela é a extração POR SS que o site guarda
    # (julho-2026.json): para cada SS de julho, as ocorrências do próprio trafo nas três
    # colunas, com a mesma janela. Serve para o resultado pelo trafo; não serve para a chave
    # gêmea nem para a busca em observação, que precisam do arquivo inteiro.
    julho_por_ss = {}
    if a.julho_json and os.path.exists(a.julho_json):
        import json as _json
        jj = _json.load(open(a.julho_json, encoding="utf-8"))
        for r in jj.get("registros", []) + jj.get("ampliado", []):
            julho_por_ss.setdefault(r["ss"], r)
        ja = {x["ss"] for x in ss}
        novos = 0
        for r in jj.get("registros", []):          # só o recorte oficial do mês, não o entorno
            if r["ss"] in ja:
                continue
            ss.append({
                "ss": txt(r["ss"]), "os": txt(r.get("os")), "obra": txt(r.get("obra")), "trafo": txt(r.get("trafo")),
                "abertura": data(r.get("abertura")), "origem": txt(r.get("origem")), "defeito": txt(r.get("defeito")),
                "tipo_ss": txt(r.get("tipo_ss")), "situacao": txt(r.get("situacao")), "criticidade": txt(r.get("criticidade")),
                "localidade": txt(r.get("localidade")), "alimentador": txt(r.get("alimentador")), "equipe": r["ss"].split(" ")[0],
                "pot_ret": txt(r.get("pot_ret")), "pot_inst": txt(r.get("pot_inst")), "descricao": txt(r.get("texto_ss"))[:400],
                "fonte": "base @ de 11/08 via julho-2026.json",
            })
            novos += 1
        print(f"  julho-2026.json: {len(julho_por_ss)} SS com extração da Crítica de julho · {novos} SS de julho acrescentadas à base")
        fonte_ss += f" + julho-2026.json ({novos} SS de julho após o fim da V4)"
    total_base = len(ss)
    fora_prefixo = collections.Counter(s["trafo"][:2] for s in ss if s["trafo"][:2] not in PREFIXOS)
    malformados = [s for s in ss if s["trafo"][:2] in PREFIXOS and not re.fullmatch(r"\d{10}", s["trafo"])]
    ss = [s for s in ss if s["trafo"][:2] in PREFIXOS and re.fullmatch(r"\d{10}", s["trafo"])]
    fora_data = [s for s in ss if not (s["abertura"] and INICIO <= s["abertura"] < FIM)]
    ss = [s for s in ss if s["abertura"] and INICIO <= s["abertura"] < FIM]
    # uma SS por linha: a mesma SS repetida na base conta uma vez
    vistos, unicos = set(), []
    for s in ss:
        if s["ss"] not in vistos:
            vistos.add(s["ss"]); unicos.append(s)
    dup = len(ss) - len(unicos)
    ss = unicos
    print(f"  no recorte (prefixo {'/'.join(PREFIXOS)}, jan–jul): {len(ss)} · fora do prefixo {sum(fora_prefixo.values())} {dict(fora_prefixo)} · fora da data {len(fora_data)} · SS repetida {dup}")

    print("Crítica:")
    ocs, por_codigo = carregar_critica(os.path.join(PUB, "bases", "originais", "Original_Critica_Interrupcoes.zip"), a.critica_extra)
    meses = sorted({o["ini"].strftime("%Y-%m") for o in ocs.values() if o["ini"]})
    print(f"  {len(ocs)} ocorrências · meses cobertos: {', '.join(meses)}")
    ultimo = max(o["fim"] for o in ocs.values() if o["fim"])
    site = carregar_site()

    linhas = []
    for s in ss:
        cod = s["trafo"]
        chave = "03" + cod[2:]
        st, o, pap, d, n = procurar(cod, s["abertura"], ocs, por_codigo)
        via, st_chave, o_ch, pap_ch, d_ch, n_ch = "trafo", "", None, "", None, 0
        if st == "AUSENTE":
            st_chave, o_ch, pap_ch, d_ch, n_ch = procurar(chave, s["abertura"], ocs, por_codigo)
            if st_chave != "AUSENTE":
                via = "chave gêmea"
        if via == "chave gêmea":
            resultado = {"SIM": "CASOU PELA CHAVE GÊMEA", "fora da janela": "CHAVE GÊMEA FORA DA JANELA"}[st_chave]
            o_mostra, pap_mostra, d_mostra = o_ch, pap_ch, d_ch
        else:
            resultado = {"SIM": "CASOU PELO TRAFO", "fora da janela": "TRAFO FORA DA JANELA", "AUSENTE": "AUSENTE — nem trafo nem chave"}[st]
            o_mostra, pap_mostra, d_mostra = o, pap, d
        # julho sem Crítica bruta: usa a extração por SS do site, e diz isso
        cobertura = "Crítica carregada" if s["abertura"] <= ultimo else "SEM CRÍTICA DESTE PERÍODO — não conferível"
        jr = julho_por_ss.get(s["ss"])
        # SS de julho: a Crítica jan–jun só alcança quem casou numa ocorrência de junho que se
        # estende até a abertura; para o resto, quem responde é a extração de julho por SS.
        if jr and s["abertura"].month == 7 and (s["abertura"] > ultimo or st != "SIM") and not resultado.startswith("CASOU"):
            cobertura = "Crítica de julho · extração por SS de 07/08 (site) — chave gêmea e observação não conferíveis"
            st_j = txt(jr.get("critica"))
            ocs_j = jr.get("ocorrencias") or []
            dentro = [o for o in ocs_j if o.get("na_janela")]
            dentro.sort(key=lambda o: ("problema" not in txt(o.get("papeis")), abs(o.get("delta_inicio_h") or 0)))
            perto = sorted(ocs_j, key=lambda o: abs(o.get("delta_inicio_h") or 9e9))
            oj = (dentro or perto or [None])[0]
            resultado = {"SIM": "CASOU PELO TRAFO", "fora da janela": "TRAFO FORA DA JANELA",
                         "AUSENTE": "AUSENTE — nem trafo nem chave"}.get(st_j, "AUSENTE — nem trafo nem chave")
            via = "trafo" if st_j in ("SIM", "fora da janela") else "—"
            n = len(ocs_j); n_ch = None
            if oj:
                o_mostra = {"oc": "", "ini": data(oj.get("inicio")), "fim": data(oj.get("fim")), "passos": "",
                            "causa": txt(oj.get("causa")), "subcausa": txt(oj.get("subcausa")),
                            "clientes": txt(oj.get("clientes")), "duracao": txt(oj.get("duracao_min")), "obs": txt(oj.get("observacao"))}
                pap_mostra = papeis_texto(txt(oj.get("papeis")).split("+"))
                ini_j, fim_j = data(oj.get("inicio")), data(oj.get("fim"))
                d_mostra = None if oj.get("na_janela") else (round(distancia_a_janela(s["abertura"], ini_j, fim_j or ini_j), 1) if ini_j else None)
            else:
                o_mostra, pap_mostra, d_mostra = None, "", None
            if resultado.startswith("AUSENTE"):
                resultado = "AUSENTE pelo trafo — chave gêmea não conferível (Crítica bruta de julho perdida)"
        elif s["abertura"].month == 7 and st != "SIM" and not resultado.startswith("CASOU"):
            # SS de julho fora do julho-2026.json: a Crítica jan–jun não cobre a abertura dela, então
            # "fora da janela" ou "ausente" medido só em jan–jun seria conclusão falsa. Só o casamento
            # real (ocorrência de junho que se estende até a abertura) vale.
            resultado = "SEM CRÍTICA — não conferível"
            cobertura = "SEM CRÍTICA DESTE PERÍODO — não conferível"
            via = "—"
            o_mostra, pap_mostra, d_mostra = None, "", None
        r = site.get(s["ss"])
        linhas.append({
            "SS": s["ss"], "OS": s["os"], "Obra": s["obra"], "Transformador": cod, "Prefixo": cod[:2],
            "Chave gêmea (03)": chave, "Abertura da SS": s["abertura"], "Mês": s["abertura"].strftime("%m/%Y"),
            "Origem SS": s["origem"], "Defeito SS": s["defeito"], "Tipo SS": s["tipo_ss"], "Situação": s["situacao"],
            "Criticidade": s["criticidade"], "Localidade": s["localidade"], "Alimentador": s["alimentador"],
            "Equipe": s["equipe"], "kVA retirado": s["pot_ret"], "kVA instalado": s["pot_inst"],
            "Cobertura da Crítica": cobertura,
            "Resultado": resultado, "Encontrado por": via if resultado.startswith("CASOU") or "FORA" in resultado else "—",
            "Ocorrências do trafo na Crítica": n, "Ocorrências da chave na Crítica": n_ch,
            "Ocorrência": o_mostra["oc"] if o_mostra else "",
            "Início da ocorrência": o_mostra["ini"] if o_mostra else None,
            "Fim da ocorrência": o_mostra["fim"] if o_mostra else None,
            "Passos": o_mostra["passos"] if o_mostra else "",
            "Papel do ativo": pap_mostra, "Distância à janela (h)": d_mostra,
            "Causa": o_mostra["causa"] if o_mostra else "", "Subcausa": o_mostra["subcausa"] if o_mostra else "",
            "Clientes": o_mostra["clientes"] if o_mostra else "", "Duração (min)": o_mostra["duracao"] if o_mostra else "",
            "Observação da Crítica": (o_mostra["obs"] if o_mostra else "")[:300],
            "Site · cascata": txt(r.get("cascata")) if r else "não está no site",
            "Site · Crítica": txt(r.get("censo_critica")) if r else "",
            "Site · gatilho de exclusão": txt(r.get("expurgo_gatilho")) if r else "",
            "Site · confirmado": txt(r.get("confirmado")) if r else "",
            "Descrição da SS": s["descricao"],
        })

    # ── resumo ─────────────────────────────────────────────────────────────
    conf = [l for l in linhas if l["Cobertura da Crítica"].startswith("Crítica")]
    nconf = [l for l in linhas if not l["Cobertura da Crítica"].startswith("Crítica")]
    julho_site = [l for l in conf if "extração por SS" in l["Cobertura da Crítica"]]
    res = collections.Counter(l["Resultado"] for l in conf)
    por_mes = collections.defaultdict(collections.Counter)
    for l in linhas:
        por_mes[l["Mês"]][l["Resultado"] if l in conf else "não conferível (sem Crítica)"] += 1
    por_pref = collections.defaultdict(collections.Counter)
    for l in conf:
        por_pref[l["Prefixo"]][l["Resultado"]] += 1
    # o que a chave mudou em relação ao site
    chave_casou = [l for l in conf if l["Resultado"] == "CASOU PELA CHAVE GÊMEA"]
    chave_site_excl = [l for l in chave_casou if l["Site · cascata"] == "EXCLUÍDA" and l["Site · gatilho de exclusão"] in ("sem_interrupcao", "fora_da_janela")]

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Resumo"
    neg = Font(bold=True)
    def linha(*vals, bold=False):
        ws.append(list(vals))
        if bold:
            for c in ws[ws.max_row]: c.font = neg
    linha("SS de transformador de 2026 × Crítica — pelo trafo e, se ausente, pela chave gêmea", bold=True)
    linha(f"Gerado em {dt.datetime.now():%d/%m/%Y %H:%M} · base de SS: {fonte_ss} · Crítica: {', '.join(meses)}")
    linha("")
    linha("Recorte", bold=True)
    linha("Linhas na base de SS", total_base)
    linha(f"Fora do prefixo {'/'.join(PREFIXOS)}", sum(fora_prefixo.values()), ", ".join(f"{k}={v}" for k, v in fora_prefixo.most_common()))
    linha("Prefixo certo mas código malformado (não são 10 dígitos)", len(malformados), ", ".join(s["trafo"] for s in malformados[:6]))
    linha("Fora de janeiro a julho", len(fora_data))
    linha("SS repetida na base (contada uma vez)", dup)
    linha("SS no recorte", len(linhas), bold=True)
    linha("  …das quais SS canceladas (ficam no recorte; filtre pela coluna Situação se quiser)", sum(1 for l in linhas if l["Situação"] == "SS CANCELADA"))
    linha("  …das quais SS repassadas", sum(1 for l in linhas if l["Situação"] == "SS REPASSADA"))
    linha("  com Crítica carregada para o período", len(conf))
    linha("    …das quais julho, pela extração por SS de 07/08 guardada no site", len(julho_site))
    linha("  sem Crítica do período (não conferível)", len(nconf))
    linha("")
    linha("Por tipo de SS (recorte inteiro)", bold=True)
    for k, v in collections.Counter(l["Tipo SS"] for l in linhas).most_common():
        linha(k, v)
    linha("")
    linha("Resultado — só as conferíveis", bold=True)
    for k in ("CASOU PELO TRAFO", "TRAFO FORA DA JANELA", "CASOU PELA CHAVE GÊMEA", "CHAVE GÊMEA FORA DA JANELA",
              "AUSENTE — nem trafo nem chave", "AUSENTE pelo trafo — chave gêmea não conferível (Crítica bruta de julho perdida)"):
        linha(k, res.get(k, 0), f"{res.get(k, 0) / max(1, len(conf)) * 100:.1f}%")
    linha("")
    linha("O que a chave gêmea acrescentou", bold=True)
    linha("Casaram só pela chave (trafo ausente da Crítica)", len(chave_casou))
    linha("  …e o site havia excluído por falta de interrupção", len(chave_site_excl))
    linha("Ausentes pelo trafo que a chave também não achou", res.get("AUSENTE — nem trafo nem chave", 0))
    linha("")
    linha("Por mês", bold=True)
    cats = ["CASOU PELO TRAFO", "TRAFO FORA DA JANELA", "CASOU PELA CHAVE GÊMEA", "CHAVE GÊMEA FORA DA JANELA", "AUSENTE — nem trafo nem chave", "não conferível (sem Crítica)"]
    linha("Mês", *cats, "Total", bold=True)
    for m in sorted(por_mes):
        linha(m, *[por_mes[m].get(c, 0) for c in cats], sum(por_mes[m].values()))
    linha("")
    linha("Por prefixo (conferíveis)", bold=True)
    linha("Prefixo", *cats[:5], "Total", bold=True)
    for p in sorted(por_pref):
        linha(p, *[por_pref[p].get(c, 0) for c in cats[:5]], sum(por_pref[p].values()))
    linha("")
    linha("Método", bold=True)
    for t in [
        "Três colunas da Crítica: COD_ELE_PROBLEMA, COD_ELE_INTERROMPIDO, COD_ELE_FECHADO.",
        "Ocorrência = passos com o mesmo NUM_SEQ_OPER_INIC_HDE; janela medida do primeiro passo aberto ao último fechado.",
        "Casa quando a abertura da SS cai entre (início − 1 h) e (fim + 24 h). Igual ao site.",
        "Chave gêmea = '03' + 8 últimos dígitos do trafo. Procurada SÓ quando o trafo não aparece em papel nenhum.",
        "Distância à janela: horas que a abertura da SS ficou fora da janela [início − 1 h, fim + 24 h] da ocorrência mais próxima, quando não casou.",
        "Quando mais de uma ocorrência casa, a mostrada é a em que o trafo é o elemento com problema; empatando, a de início mais próximo da abertura.",
        "Dezembro/2025 não está carregado: SS dos primeiros dias de janeiro podem ter ocorrência lá (o site tratou 24 casos assim).",
        "Julho: a Crítica bruta de julho (Critica__072026.txt) se perdeu nos reinícios do contêiner. O resultado pelo trafo vem da extração por SS de 07/08 guardada no site (julho-2026.json), mesmo método; a chave gêmea e a busca em observação de julho ficam sem conferência até o arquivo ser reenviado.",
        "Isto é leitura ao lado do caso. Não mexe no 1.305 nem no 1.582; a coluna 'Site' mostra o que o site decidiu.",
    ]:
        linha(t)
    ws.column_dimensions["A"].width = 58
    for c in "BCDEFGH": ws.column_dimensions[c].width = 22

    def aba(nome, dados):
        w = wb.create_sheet(nome)
        if not dados:
            w.append(["nenhuma"]); return
        cols = list(dados[0].keys())
        w.append(cols)
        for c in w[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F3864")
            c.alignment = Alignment(wrap_text=True, vertical="center")
        for d in dados:
            w.append([d[c] for c in cols])
        for i, c in enumerate(cols, 1):
            w.column_dimensions[get_column_letter(i)].width = min(48, max(12, len(c) + 2))
        for row in w.iter_rows(min_row=2):
            for c in row:
                if isinstance(c.value, dt.datetime):
                    c.number_format = "dd/mm/yyyy hh:mm"
        w.freeze_panes = "B2"
        w.auto_filter.ref = w.dimensions

    aba("Todas as SS", linhas)
    aba("Casaram pela chave", chave_casou)
    aba("Ausentes", [l for l in conf if l["Resultado"].startswith("AUSENTE")])
    aba("Fora da janela", [l for l in conf if "FORA DA JANELA" in l["Resultado"]])
    aba("Não conferíveis", nconf)
    aba("Julho pela extração do site", julho_site)
    aba("Divergem do site", [l for l in conf if (l["Resultado"].startswith("CASOU") and l["Site · cascata"] == "EXCLUÍDA" and l["Site · gatilho de exclusão"] in ("sem_interrupcao", "fora_da_janela"))
                            or (l["Resultado"].startswith("AUSENTE") and l["Site · Crítica"] == "DEFEITO NA JANELA")])
    os.makedirs(os.path.dirname(os.path.abspath(a.saida)), exist_ok=True)
    wb.save(a.saida)
    print(f"\n{a.saida}")
    print(f"SS no recorte {len(linhas)} · conferíveis {len(conf)} · não conferíveis {len(nconf)}")
    for k in cats[:5]:
        print(f"  {k}: {res.get(k, 0)}")
    print(f"  chave casou e o site excluía por falta de interrupção: {len(chave_site_excl)}")


if __name__ == "__main__":
    principal()
