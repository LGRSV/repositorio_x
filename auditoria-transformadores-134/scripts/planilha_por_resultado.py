"""Uma aba por resultado, com a verificação pela OBSERVACAO já dentro de cada linha.

Junta a saída de analise_ss_critica_chave.py (uma linha por SS, com o resultado pelo trafo
e pela chave gêmea) com a de ausentes_na_observacao.py (o que o texto de observação da
Crítica diz sobre os ausentes e os fora da janela), e grava um Excel com:
  Resumo · Casou pelo trafo · Trafo fora da janela · Casou pela chave gêmea ·
  Chave fora da janela · Ausente de tudo · Julho sem Crítica
"""

import argparse
import collections
import datetime as dt
import json
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ABAS = [
    ("CASOU PELO TRAFO", "Casou pelo trafo"),
    ("TRAFO FORA DA JANELA", "Trafo fora da janela"),
    ("CASOU PELA CHAVE GÊMEA", "Casou pela chave gêmea"),
    ("CHAVE GÊMEA FORA DA JANELA", "Chave fora da janela"),
    ("AUSENTE — nem trafo nem chave", "Ausente de tudo"),
    ("AUSENTE pelo trafo — chave gêmea não conferível (Crítica bruta de julho perdida)", "Julho · ausente pelo trafo"),
    ("SEM CRÍTICA", "Sem Crítica do período"),
]
OBS_COLS = ["Observação · veredito", "Observação · menções", "Observação · na janela",
            "Observação · ocorrências na janela", "Observação · elemento com problema", "Observação · trecho",
            "Observação · leitura do revisor"]


def principal():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ss", required=True, help="SS_x_Critica_trafo_e_chave_2026.xlsx")
    ap.add_argument("--obs", required=True, help="Ausentes_na_Observacao_da_Critica*.json")
    ap.add_argument("--saida", required=True)
    ap.add_argument("--leitura", default=None, help="JSON SS → leitura humana/agente do trecho (opcional)")
    a = ap.parse_args()

    wb0 = openpyxl.load_workbook(a.ss, read_only=True)
    ws0 = wb0["Todas as SS"]
    rows = ws0.iter_rows(values_only=True)
    cols = list(next(rows))
    linhas = [dict(zip(cols, r)) for r in rows]

    obs = json.load(open(a.obs, encoding="utf-8"))
    por_ss = {r["SS"]: r for r in obs["resumo"]}
    leitura = json.load(open(a.leitura, encoding="utf-8")) if a.leitura else {}

    for l in linhas:
        o = por_ss.get(l["SS"])
        if o:
            l["Observação · veredito"] = o["Veredito"]
            l["Observação · menções"] = o["Menções na observação"]
            l["Observação · na janela"] = o["…na janela"]
            l["Observação · ocorrências na janela"] = o["Ocorrências (na janela)"]
            l["Observação · elemento com problema"] = o["Elemento com problema (na janela)"]
            l["Observação · trecho"] = o["Primeiro trecho"]
            l["Observação · leitura do revisor"] = leitura.get(l["SS"], "")
        else:
            for c in OBS_COLS:
                l[c] = "" if c not in ("Observação · menções", "Observação · na janela") else None
            l["Observação · veredito"] = "não se aplica (casou nas colunas de ativo)" if str(l["Resultado"]).startswith("CASOU") else ""
        julho_site = "extração por SS" in str(l["Cobertura da Crítica"])
        if julho_site:
            # a observação de julho não pôde ser varrida: a Crítica bruta do mês se perdeu
            l["Observação · veredito"] = "não conferível — Crítica bruta de julho perdida" if not str(l["Resultado"]).startswith("CASOU") else "não se aplica (casou nas colunas de ativo)"
            for c in OBS_COLS[1:]:
                l[c] = None if c in ("Observação · menções", "Observação · na janela") else ""
        l["Grupo"] = "SEM CRÍTICA" if not str(l["Cobertura da Crítica"]).startswith("Crítica") else l["Resultado"]

    ordem = ["SS", "OS", "Obra", "Transformador", "Prefixo", "Chave gêmea (03)", "Abertura da SS", "Mês",
             "Origem SS", "Defeito SS", "Tipo SS", "Situação", "Criticidade", "Localidade", "Alimentador", "Equipe",
             "kVA retirado", "kVA instalado", "Resultado", "Encontrado por", "Cobertura da Crítica"] + OBS_COLS + [
             "Ocorrências do trafo na Crítica", "Ocorrências da chave na Crítica", "Ocorrência", "Início da ocorrência",
             "Fim da ocorrência", "Passos", "Papel do ativo", "Distância à janela (h)", "Causa", "Subcausa", "Clientes",
             "Duração (min)", "Observação da Crítica", "Site · cascata", "Site · Crítica", "Site · gatilho de exclusão",
             "Site · confirmado", "Descrição da SS"]

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Resumo"
    ws.append(["SS de transformador (42/52/53/57, jan–jul/2026) × Crítica — uma aba por resultado"]); ws["A1"].font = Font(bold=True)
    ws.append([f"Gerado em {dt.datetime.now():%d/%m/%Y %H:%M}"])
    ws.append([])
    ws.append(["Resultado", "SS", "…trafo mencionado na observação, na janela", "…mencionado fora da janela", "…não mencionado", "…só outro equipamento de mesmo final"])
    for c in ws[4]: c.font = Font(bold=True)
    grupos = collections.defaultdict(list)
    for l in linhas:
        grupos[l["Grupo"]].append(l)
    for chave, nome in ABAS:
        g = grupos.get(chave, [])
        v = collections.Counter(l["Observação · veredito"] for l in g)
        ws.append([nome, len(g), v.get("MENCIONADO NA JANELA", 0), v.get("MENCIONADO FORA DA JANELA", 0), v.get("NÃO MENCIONADO", 0), v.get("SÓ OUTRO EQUIPAMENTO DE MESMO FINAL", 0)])
    ws.append(["Total", len(linhas)])
    ws.append([])
    jul = [l for l in linhas if "extração por SS" in str(l["Cobertura da Crítica"])]
    ws.append(["Julho pela extração por SS de 07/08 guardada no site (dentro dos totais acima)", len(jul)]); ws[ws.max_row][0].font = Font(bold=True)
    for chave, nome in ABAS:
        n = sum(1 for l in jul if l["Grupo"] == chave)
        if n:
            ws.append([f"  {nome}", n])
    ws.append([])
    for t in ["Método: três colunas de ativo da Crítica; ocorrência do primeiro passo ao último; casa se a abertura da SS está entre (início − 1h) e (fim + 24h).",
              "Chave gêmea = 03 + 8 dígitos finais, procurada só quando o trafo não aparece em papel nenhum.",
              "Observação: para ausentes e fora da janela, o código do trafo (inteiro ou 8 finais com qualquer prefixo) e o número da SS com a sigla da equipe foram procurados no texto OBSERVACAO de todos os passos.",
              "Mencionado na observação não é o mesmo que ter defeito registrado: veja o elemento com problema, o trecho e a leitura do revisor na linha.",
              "Os 8 dígitos finais com prefixo 79/03/02/67/33/88 apontam para religador ou chave do mesmo ponto, não para o trafo — contados à parte como 'outro equipamento'.",
              "Número da SS só conta com a sigla da equipe junto (ENC-RD-PS 26 não é ETO-RD-AG 26).",
              "Dezembro/2025 não carregado (SS dos primeiros dias de janeiro podem ter ocorrência lá).",
              "Julho: a Crítica bruta (Critica__072026.txt) se perdeu nos reinícios do contêiner. O resultado pelo trafo vem da extração por SS de 07/08 guardada no site, mesmo método e mesma janela; a chave gêmea e a observação de julho ficam sem conferência até o arquivo ser reenviado.",
              "Leitura ao lado do caso. Não mexe no 1.305 nem no 1.582; as colunas 'Site' mostram o que o site decidiu."]:
        ws.append([t])
    ws.column_dimensions["A"].width = 40
    for c in "BCDE": ws.column_dimensions[c].width = 26

    def aba(nome, dados):
        w = wb.create_sheet(nome)
        w.append(ordem)
        for c in w[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F3864"); c.alignment = Alignment(wrap_text=True, vertical="center")
        for d in dados:
            w.append([d.get(c) for c in ordem])
        for i, c in enumerate(ordem, 1):
            w.column_dimensions[get_column_letter(i)].width = min(50, max(12, len(c) + 2))
        for row in w.iter_rows(min_row=2):
            for c in row:
                if isinstance(c.value, dt.datetime):
                    c.number_format = "dd/mm/yyyy hh:mm"
        w.freeze_panes = "B2"
        if dados:
            w.auto_filter.ref = w.dimensions

    for chave, nome in ABAS:
        if chave == "SEM CRÍTICA" and not grupos.get(chave):
            continue
        aba(nome, sorted(grupos.get(chave, []), key=lambda l: l["Abertura da SS"]))
    aba("Julho (detalhe)", sorted(jul, key=lambda l: l["Abertura da SS"]))
    wb.save(a.saida)
    print(a.saida)
    for chave, nome in ABAS:
        print(f"  {nome}: {len(grupos.get(chave, []))}")


if __name__ == "__main__":
    principal()
