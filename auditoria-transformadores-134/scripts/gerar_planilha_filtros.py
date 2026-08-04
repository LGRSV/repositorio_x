#!/usr/bin/env python3
"""Gera Filtros_do_Site.xlsx — todos os filtros do site, aba por aba, com o cruzamento dinâmico.

Pedido dele: "um excel com todos os filtros que nem no site, com as informações por aba e um
resumo dinâmico final com o site".

A composição de cada filtro NÃO é recalculada aqui. Reimplementar em Python as regras que a tela
aplica seria criar uma segunda verdade, e duas verdades divergem no dia em que ninguém está
olhando. Em vez disso, um robô abre o site, clica cada filtro e baixa a planilha dele — o que
entra neste arquivo é o que a tela mostra, porque veio da tela. O extrator é
scripts/../extrai_filtros.mjs e ele grava filtros.json; este script só monta o Excel.

Abas:
  · Resumo            — a conta inteira, na ordem da Visão geral
  · Filtros por aba   — cada filtro do site com quantos casos tem e o que ele significa
  · Filtro × SS       — tabela longa (aba, filtro, SS): é daqui que sai qualquer tabela dinâmica
  · Dimensões         — uma linha por SS com os campos pelos quais o site filtra
  · Uma aba por tela  — os filtros daquela tela, na ordem em que aparecem

Como cruzar: em "Filtro × SS" faça a dinâmica que quiser; para trazer qualquer atributo do caso,
PROCV da coluna SS contra a aba Dimensões.

Uso:  python3 gerar_planilha_filtros.py [caminho/para/filtros.json]
"""
import datetime
import json
import os
import re
import sys
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FLUXO = os.path.join(RAIZ, "public", "fluxo-1510.json")
SAIDA = os.path.join(RAIZ, "public", "bases", "Filtros_do_Site.xlsx")
PADRAO_FILTROS = os.path.join(RAIZ, "public", "filtros-do-site.json")

TINTA = PatternFill("solid", fgColor="1f2a37")
CLARO = PatternFill("solid", fgColor="eef2f7")

# Os campos pelos quais o site filtra, com o nome que ele usa na tela. É esta a aba que permite
# refazer qualquer número do site numa tabela dinâmica sem depender dos filtros prontos.
DIMENSOES = [
    ("SS", "ss"), ("OS", "os"), ("Obra", "obra"), ("Transformador", "trafo"),
    ("Localidade", "localidade"), ("Alimentador", "alimentador"), ("Origem", "origem"),
    ("Solicitante", "solicitante"), ("Equipe da SS", "equipe_ss"),
    ("Abertura", "abertura"), ("Término", "termino"), ("Situação", "situacao"),
    ("Onde está hoje", "cascata"), ("Motivo", "cascata_motivo"), ("Decisão", "decisao"),
    ("Causa confirmada", "confirmado"), ("Categoria de exclusão", "expurgo_gatilho"),
    ("Por que saiu", "exclusao_porque"), ("Etiquetas", "etiquetas"),
    ("Fato", "fato"), ("Leitura", "leitura"), ("Categoria pelo texto", "categoria_texto"),
    ("Categoria gravada", "categoria_gravada"), ("Censo da Crítica", "censo_critica"),
    ("Nível da interrupção", "e1_nivel"), ("Elemento do defeito", "def_elemento"),
    ("Ocorrência", "oc_num"), ("Início da ocorrência", "oc_ini"), ("Fim da ocorrência", "oc_fim"),
    ("Duração (h)", "oc_dur_h"), ("Distância até o intervalo (h)", "oc_dist_h"),
    ("Clientes interrompidos", "oc_cons"), ("Causa em campo", "oc_causa"),
    ("Subcausa em campo", "oc_sub"), ("Contida na SS", "oc_contida_na_ss"),
    ("Aberta antes", "aberta_antes"), ("Ressalvas", "ressalvas"),
    ("Deslocamento", "deslocamento"), ("Atendimento", "at_num"), ("Causa do atendimento", "at_causa"),
    ("Subcausa do atendimento", "at_sub"), ("Equipe do atendimento", "at_equipe"),
    ("Material conferido", "material_conferido"), ("Trafos no material", "trafos_material"),
    ("Só falta o SIAGO", "pendente_siago"), ("Status da obra", "obra_status"),
    ("Tipo da obra", "obra_tipo"), ("Descrição da obra", "obra_descricao"),
    ("Empreiteira", "obra_empreiteira"), ("SIGCO", "sigco"),
    ("Potência retirada", "pot_ret"), ("Potência instalada", "pot_inst"),
    ("Sem cliente interrompido", "sem_cliente_interrompido"), ("Sob suspeita", "sob_suspeita"),
    ("Proteção que atuou", "protecao"), ("Idade da SS (dias)", "idade_ss"),
]


def t(v):
    return "" if v is None else str(v).strip()


def cabecalho(ws, colunas, nota="", largura=None):
    if nota:
        ws.append([nota])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(colunas)))
        c = ws.cell(row=1, column=1)
        c.font = Font(italic=True, color="44506a")
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[1].height = 40
        ws.append([])
    ini = ws.max_row + 1
    ws.append(colunas)
    for c in ws[ini]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = TINTA
    ws.freeze_panes = ws.cell(row=ini + 1, column=1)
    if largura:
        for j, w in enumerate(largura, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    return ini


def autolargura(ws, ini, maximo=70, amostra=300):
    largos = {}
    for i, linha in enumerate(ws.iter_rows(min_row=ini, values_only=True)):
        if i > amostra:
            break
        for j, v in enumerate(linha, start=1):
            largos[j] = max(largos.get(j, 10), len(str(v)) if v is not None else 0)
    for j, w in largos.items():
        ws.column_dimensions[get_column_letter(j)].width = min(maximo, max(11, w + 2))


def main(argv):
    caminho = argv[0] if argv else PADRAO_FILTROS
    if not os.path.exists(caminho):
        print(f"falta {caminho} — rode antes o extrator extrai_filtros.mjs")
        return 2
    with open(caminho, encoding="utf-8") as fh:
        filtros = json.load(fh)
    with open(FLUXO, encoding="utf-8") as fh:
        fluxo = json.load(fh)
    R = fluxo["registros"]
    porss = {t(r.get("ss")): r for r in R}
    carimbo = fluxo["meta"].get("revisado_em") or str(datetime.date.today())

    # um filtro com 0 casos continua sendo filtro: some da tabela longa, não da lista
    filtros = [f for f in filtros if t(f.get("filtro"))]
    divergem = [f for f in filtros if f["quantos"] != f["baixados"]]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ------------------------------------------------------------------ resumo
    ws = wb.create_sheet("Resumo")
    PAROU = {"fora_da_janela", "sem_interrupcao", "sem_fato"}
    exc = [r for r in R if t(r.get("cascata")) == "EXCLUÍDA"]
    critica = [r for r in exc if t(r.get("expurgo_gatilho")) in PAROU]
    pos = [r for r in R if r not in critica]
    outra = [r for r in pos if t(r.get("cascata")) == "EXCLUÍDA"]
    retido = [r for r in pos if t(r.get("cascata")).startswith("RETIDO")]
    saida = [r for r in R if t(r.get("cascata")) == "SAÍDA"]
    ini = cabecalho(ws, ["", "Quantas", "O que é"],
                    f"A conta inteira, na mesma ordem da Visão geral do site — gerada de "
                    f"fluxo-1510.json em {carimbo}. Cada linha desta cascata tem filtro "
                    f"correspondente nas abas seguintes.", [46, 11, 92])
    for rot, n, nota in [
        ("Entram", len(R), "solicitações de troca de transformador, janeiro a junho de 2026"),
        ("− presas pela Crítica", len(critica), "sem interrupção que sustente o caso"),
        ("    · defeito nele, em outra data", sum(1 for r in critica if t(r.get("expurgo_gatilho")) == "fora_da_janela"), ""),
        ("    · ausente da Crítica", sum(1 for r in critica if t(r.get("expurgo_gatilho")) == "sem_interrupcao"), ""),
        ("    · sem rastro em base alguma", sum(1 for r in critica if t(r.get("expurgo_gatilho")) == "sem_fato"), ""),
        ("= passaram pela Crítica", len(pos), "é sobre este conjunto que as perguntas seguintes são feitas"),
        ("    · sem deslocamento do TMAE", sum(1 for r in pos if t(r.get("deslocamento")) == "SEM REGISTRO"),
         "marcador — sinaliza, não subtrai"),
        ("− excluídas de fato", len(outra), "outra causa declarada no texto ou na obra"),
        ("− sem prova de troca", len(retido), "a obra não comprova que um transformador saiu do poste"),
        ("    · só falta o SIAGO", sum(1 for r in retido if t(r.get("pendente_siago")) == "SIM"), ""),
        ("= queimados e avariados", len(saida), "o número do indicador"),
        ("    · queimados", sum(1 for r in saida if t(r.get("confirmado")) == "QUEIMADO"), ""),
        ("    · avariados", sum(1 for r in saida if t(r.get("confirmado")) == "AVARIADO"), ""),
    ]:
        ws.append([rot, n, nota])
        if not rot.startswith("    "):
            for c in ws[ws.max_row]:
                c.font = Font(bold=True)
    ws.append([])
    ws.append(["Filtros do site nesta planilha", len(filtros), "clicados um a um na tela"])
    ws.append(["Casos somados em todos os filtros", sum(f["baixados"] for f in filtros),
               "um mesmo caso aparece em vários filtros — é assim que a tela funciona"])
    ws.append(["Filtros em que a contagem da tela diverge da planilha baixada", len(divergem),
               "zero é o esperado; qualquer número aqui é defeito a investigar"])
    for c in ws[ws.max_row]:
        c.fill = CLARO

    # --------------------------------------------------------- filtros por aba
    ws = wb.create_sheet("Filtros por aba")
    ini = cabecalho(ws, ["Aba do site", "Filtro", "Casos", "O que este filtro significa"],
                    "Todos os filtros do site, na ordem em que aparecem em cada tela. A coluna "
                    "Casos é o que a tela mostra ao clicar — não é recalculada aqui.",
                    [30, 52, 9, 110])
    for f in filtros:
        ws.append([f["aba"], f["filtro"], f["quantos"], f["nota"]])
    for linha in ws.iter_rows(min_row=ini + 1):
        linha[3].alignment = Alignment(wrap_text=True, vertical="top")
    ws.auto_filter.ref = f"A{ini}:D{ws.max_row}"

    # ------------------------------------------------------------- filtro × SS
    ws = wb.create_sheet("Filtro x SS")
    ini = cabecalho(ws, ["Aba do site", "Filtro", "SS", "Onde está hoje", "Causa confirmada",
                         "Categoria de exclusão", "Localidade", "Mês de abertura"],
                    "A tabela longa: uma linha por par filtro/solicitação. É daqui que sai "
                    "qualquer tabela dinâmica — conte SS por filtro, cruze filtro com localidade, "
                    "com mês, com o destino. Para outros atributos, PROCV da coluna SS contra a "
                    "aba Dimensões.", [28, 46, 22, 30, 18, 26, 22, 15])
    for f in filtros:
        for ss in f["ss"]:
            r = porss.get(ss, {})
            ws.append([f["aba"], f["filtro"], ss, t(r.get("cascata")), t(r.get("confirmado")),
                       t(r.get("expurgo_gatilho")), t(r.get("localidade")),
                       t(r.get("abertura"))[:7]])
    ws.auto_filter.ref = f"A{ini}:H{ws.max_row}"

    # ------------------------------------------------------------- dimensões
    ws = wb.create_sheet("Dimensões")
    ini = cabecalho(ws, [n for n, _ in DIMENSOES],
                    "Uma linha por solicitação com os campos pelos quais o site filtra. Serve "
                    "para duas coisas: refazer qualquer número do site numa dinâmica própria, e "
                    "trazer atributos para a aba Filtro x SS com PROCV pela coluna SS.")
    for r in R:
        ws.append([t(r.get(c)) for _, c in DIMENSOES])
    autolargura(ws, ini, maximo=42)
    ws.auto_filter.ref = f"A{ini}:{get_column_letter(len(DIMENSOES))}{ws.max_row}"

    # ------------------------------------------------------- uma aba por tela
    por_aba = defaultdict(list)
    for f in filtros:
        por_aba[f["aba"]].append(f)
    usados = set()
    for aba, fs in por_aba.items():
        # o nome da aba do Excel não aceita tudo o que o menu do site aceita
        titulo = re.sub(r"[\\/*?:\[\]]", "-", aba)[:31] or "aba"
        while titulo in usados:
            titulo = titulo[:28] + "_" + str(len(usados))
        usados.add(titulo)
        ws = wb.create_sheet(titulo)
        ini = cabecalho(ws, ["Filtro", "Casos", "% do recorte", "O que este filtro significa"],
                        f"Tela “{aba}” do site — {len(fs)} filtros, na ordem em que "
                        f"aparecem. Percentual sobre as {len(R)} solicitações do recorte.",
                        [54, 9, 13, 110])
        for f in fs:
            ws.append([f["filtro"], f["quantos"], round(100 * f["quantos"] / len(R), 1), f["nota"]])
        for linha in ws.iter_rows(min_row=ini + 1):
            linha[3].alignment = Alignment(wrap_text=True, vertical="top")

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    print(SAIDA)
    print(f"  {len(filtros)} filtros em {len(por_aba)} telas · "
          f"{sum(f['baixados'] for f in filtros)} linhas em Filtro x SS · "
          f"{len(R)} linhas em Dimensões")
    if divergem:
        print(f"  ATENÇÃO: {len(divergem)} filtros com contagem divergente da planilha baixada:")
        for f in divergem[:10]:
            print(f"    {f['aba']} · {f['filtro']}: tela {f['quantos']}, planilha {f['baixados']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
