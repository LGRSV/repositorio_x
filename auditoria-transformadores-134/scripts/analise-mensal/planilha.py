"""Gera a planilha da análise mensal a partir do resultado de rodar_mes.py."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AZUL = PatternFill("solid", fgColor="1F3864")
FINO = Side(style="thin", color="BFBFBF")
BORDA = Border(left=FINO, right=FINO, top=FINO, bottom=FINO)
COR = {"SIM": "C6E0B4", "fora da janela": "FFF2CC", "AUSENTE": "FFC7CE"}


def _cabeca(ws, colunas, altura=42):
    ws.append(colunas)
    for c in ws[1]:
        c.fill = AZUL
        c.font = Font(color="FFFFFF", bold=True, size=9)
        c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    ws.row_dimensions[1].height = altura
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}{max(2, ws.max_row)}"


def gerar(linhas, caminho, mes, arquivo_critica):
    wb = Workbook()

    # ── Leia-me ─────────────────────────────────────────────────────────────
    rs = wb.active
    rs.title = "Leia-me"

    def bl(t, b=False, sz=10, cor="000000"):
        rs.append([t])
        rs.cell(rs.max_row, 1).font = Font(bold=b, size=sz, color=cor)
        rs.cell(rs.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")

    bl(f"Análise de {mes}", True, 14, "1F3864")
    bl(f"Gerada por scripts/analise-mensal/rodar_mes.py · Crítica usada: {arquivo_critica}")
    bl("")
    bl("O QUE ESTE ARQUIVO É E O QUE NÃO É", True, 12, "1F3864")
    bl("É o que as bases sabem dizer sozinhas. A coluna ENTRA só diz SIM quando o caso "
       "casa na Crítica dentro da janela, não tem suspeita de auxiliar e não tem nenhuma "
       "pendência. Todo o resto sai como PENDENTE, com o motivo escrito.")
    bl("O indicador do mês NÃO sai daqui pronto. As pendências dependem de informação "
       "que a base não tem: se foi cola e fita, se precisou substituir, se o auxiliar "
       "está mesmo colado no religador. Isso é martelo do dono, e a planilha existe "
       "para ele bater com o caso na frente.")
    bl("")
    bl("A JANELA", True, 12, "1F3864")
    bl("A SS casa se a abertura dela cair entre (início da ocorrência menos 1 hora) e "
       "(fim da ocorrência mais 24 horas). Só uma hora para trás.")
    bl("Datas usadas: DTA_ABERT e DTA_FECH. As colunas DTA_INIC_ELE e DTA_FIM_ELE vêm "
       "vazias em algumas extrações — quem usa elas conclui 'sem ocorrência' para todo "
       "mundo. O carregador valida isso e recusa a extração se estiver assim.")
    bl("O ativo casa em qualquer um dos três papéis: defeito, interrompido ou fechado.")
    bl("")
    bl("AUSENTE DA CRÍTICA NÃO É CONTRAPROVA", True, 12, "1F3864")
    bl("A ocorrência só entra na Crítica quando FINALIZA. Caso recente aparece como "
       "ausente e passa a casar na extração seguinte — aconteceu com três casos de "
       "agosto de 2026 entre 10/08 e 11/08. Por isso vale reprocessar o mês a cada "
       "extração nova, até o mês fechar.")
    bl("")
    bl("TRANSFORMADOR AUXILIAR", True, 12, "1F3864")
    bl("Prefixo 51 ou 57, alimentando religador (79) ou regulador (58), zero cliente, "
       "e o código costuma espelhar o do equipamento trocando só o prefixo. Não é carga "
       "de cliente e fica fora do indicador.")
    bl("14% dos auxiliares não têm código espelhado, então a ausência de espelho não "
       "descarta. A confirmação é a distância no KML: os 712 auxiliares medidos estavam "
       "todos a menos de 20 m do equipamento, e o caso mais próximo do indicador estava "
       "a 89 m.")
    bl("")
    bl("SS GÊMEA DE REPASSE", True, 12, "1F3864")
    bl("Duas SS com o mesmo timestamp ao segundo e o mesmo texto, uma REPASSADA sem OS "
       "e outra ATENDIDA com OS, são o mesmo evento. Contar as duas infla o indicador.")
    rs.column_dimensions["A"].width = 150

    # ── os casos ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Casos")
    cols = ["ENTRA?", "pendências", "SS", "transformador", "abertura", "situação", "OS",
            "origem", "defeito", "tipo de SS", "criticidade", "Crítica", "auxiliar?",
            "evidências de auxiliar", "SS gêmea", "clientes", "kVA ret", "kVA inst",
            "série retirada", "série instalada", "fabricante retirado", "obra",
            "alimentador", "localidade", "capturado por", "TEXTO DA SS", "TEXTO DA OS"]
    _cabeca(ws, cols)
    for x in linhas:
        ws.append([
            x["entra"], " · ".join(x["pendencias"]) or "—", x["ss"], x["trafo"],
            x["abertura"], x["situacao"], x["os"] or "—", x["origem"] or "—",
            x["defeito"] or "—", x["tipo_ss"], x["criticidade"], x["critica"],
            x["auxiliar"], " · ".join(x["auxiliar_evidencias"]) or "—",
            "; ".join(f"{g['ss']} ({'repasse' if g['duplicata_de_repasse'] else 'outra'})"
                      for g in x["gemeas"]) or "—",
            x["clientes"], x["pot_ret"] or "—", x["pot_inst"] or "—",
            x["ns_ret"] or "—", x["ns_inst"] or "—", x["fab_ret"] or "—",
            x["obra"] or "—", x["alimentador"], x["localidade"],
            " + ".join(x["capturado_por"]), x["texto_ss"], x["texto_os"] or "(sem texto)",
        ])
        r = ws.max_row
        fill = PatternFill("solid", fgColor=COR.get(x["critica"], "FFFFFF"))
        for c in ws[r]:
            c.fill = fill
            c.border = BORDA
            c.font = Font(size=9)
            c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(r, 1).font = Font(size=9, bold=True,
                                  color="1F3864" if x["entra"] == "SIM" else "833C00")
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="top")
        if x["auxiliar"] != "NAO":
            ws.cell(r, 13).font = Font(size=9, bold=True, color="C00000")
    larguras = {1: 11, 2: 46, 3: 22, 4: 13, 5: 18, 6: 14, 7: 22, 8: 12, 9: 22, 10: 24,
                11: 13, 12: 14, 13: 11, 14: 40, 15: 26, 16: 9, 17: 9, 18: 9, 19: 14,
                20: 14, 21: 16, 22: 13, 23: 13, 24: 18, 25: 30, 26: 60, 27: 60}
    for i, v in larguras.items():
        ws.column_dimensions[get_column_letter(i)].width = v

    # ── as ocorrências ──────────────────────────────────────────────────────
    wo = wb.create_sheet("Ocorrências da Crítica")
    _cabeca(wo, ["SS", "transformador", "início", "fim", "papel do ativo", "na janela?",
                 "Δ SS − início (h)", "Δ SS − fim (h)", "causa", "subcausa", "clientes",
                 "duração (min)", "observação"])
    for x in linhas:
        for o in x["ocorrencias"]:
            wo.append([x["ss"], x["trafo"], o["inicio"], o["fim"], o["papeis"],
                       "SIM" if o["na_janela"] else "não", o["delta_inicio_h"],
                       o["delta_fim_h"], o["causa"], o["subcausa"], o["clientes"],
                       o["duracao_min"], o["observacao"]])
            r = wo.max_row
            fill = PatternFill("solid", fgColor="C6E0B4" if o["na_janela"] else "FFF2CC")
            for c in wo[r]:
                c.fill = fill
                c.border = BORDA
                c.font = Font(size=9)
                c.alignment = Alignment(vertical="top", wrap_text=True)
    for i, v in {1: 22, 2: 13, 3: 17, 4: 17, 5: 30, 6: 11, 7: 16, 8: 16, 9: 26, 10: 30,
                 11: 9, 12: 12, 13: 80}.items():
        wo.column_dimensions[get_column_letter(i)].width = v

    wb.save(caminho)
