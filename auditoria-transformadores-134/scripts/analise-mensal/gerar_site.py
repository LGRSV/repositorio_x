"""Gera o JSON que a aba do mês lê no site.

    python3 gerar_site.py 08/2026 --pasta /bases --decisoes decisoes_agosto.json \
        --saida ../../public/agosto-2026.json

O arquivo de decisões é opcional e traz o martelo do dono por SS:

    {"ETO-RD-AG 00800/2026": {"categoria": "...", "entra": "SIM",
                              "justificativa": "...", "quem": "...", "quando": "..."}}

Sem ele, tudo sai como o script apurou, e o que depender de campo fica PENDENTE.

O JSON gerado é de um MÊS SÓ e não encosta em fluxo-1510.json. Cada mês é um
indicador separado, por decisão do dono.
"""

import argparse
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from base import ler_ss_os, ler_critica, indexar, data, veredito_critica, suspeita_auxiliar, gemeas
from rodar_mes import e_de_transformador, FALHA

MES_NOME = {"01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
            "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
            "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"}


def main():
    p = argparse.ArgumentParser()
    p.add_argument("mes")
    p.add_argument("--pasta", required=True)
    p.add_argument("--decisoes", default=None)
    p.add_argument("--universo", default=None,
                   help="xlsx do infotrafo: define o recorte oficial do mês")
    p.add_argument("--saida", required=True)
    a = p.parse_args()
    mm, aaaa = a.mes.split("/")

    cab, regs = ler_ss_os(f"{a.pasta}/BASE_SS_OS_parte1.txt", f"{a.pasta}/BASE_SS_OS_parte2.txt")
    campo = indexar(cab)
    crit = next(f"{a.pasta}/{n}" for n in os.listdir(a.pasta)
                if n.lower().startswith("critica") and n.endswith(".txt"))
    I, oc = ler_critica(crit)

    decisoes = json.load(open(a.decisoes, encoding="utf-8")) if a.decisoes else {}

    # o recorte oficial do mês: o infotrafo, quando existir
    oficial = None
    if a.universo:
        from openpyxl import load_workbook
        ws = load_workbook(a.universo, read_only=True, data_only=True)["Export"]
        it = ws.iter_rows(values_only=True)
        h = [("" if v is None else str(v)).strip() for v in next(it)]
        n = len(h)
        j = {c: i for i, c in enumerate(h)}
        oficial = set()
        for r in it:
            if not r or r[0] is None:
                continue
            r = list(r) + [""] * (n - len(r))
            if str(r[j["DATA_ABERTURA_SS"]] or "").startswith(f"{aaaa}-{mm}"):
                oficial.add(str(r[j["NUMERO_SS"]]).strip())

    do_mes = [r for r in regs if campo(r, "DATA_ABERTURA_SS")[3:10] == f"{mm}/{aaaa}"]
    trafo = [(r, e_de_transformador(campo, r)) for r in do_mes]
    trafo = [(r, q) for r, q in trafo if q]

    registros, ampliado = [], []
    for r, porques in trafo:
        ss = campo(r, "NUMERO_SS")
        cod = campo(r, "NUM_TRAFO")
        ab = data(campo(r, "DATA_ABERTURA_SS"))
        estado, detalhe = veredito_critica(ab, I, oc, cod)
        grau, ev = suspeita_auxiliar(cod, campo(r, "DESCRIPTION_SS"), campo(r, "QTD_CLIENTES"))
        gem = gemeas(campo, regs, ss)
        d = decisoes.get(ss, {})
        item = {
            "ss": ss, "trafo": cod, "abertura": campo(r, "DATA_ABERTURA_SS"),
            "situacao": campo(r, "SITUACAO_SS"), "os": campo(r, "NUMERO_OS"),
            "origem": campo(r, "ORIGEM_SS"), "defeito": campo(r, "DEFEITO_SS"),
            "tipo_ss": campo(r, "TIPOSS"), "criticidade": campo(r, "CRITICIDADE_SS"),
            "obra": campo(r, "NUM_OBRA"), "alimentador": campo(r, "ALIMENTADOR"),
            "localidade": campo(r, "LOCALIDADE"), "clientes": campo(r, "QTD_CLIENTES"),
            "pot_ret": campo(r, "POTENCIA_RET"), "pot_inst": campo(r, "POTENCIA_INST"),
            "critica": estado, "ocorrencias": detalhe,
            "auxiliar": grau, "auxiliar_evidencias": ev,
            "gemeas": gem, "capturado_por": porques,
            "texto_ss": campo(r, "DESCRIPTION_SS"), "texto_os": campo(r, "DESCRICAO_OS"),
            "no_recorte": (oficial is None or ss in oficial),
            # o martelo do dono, quando existe, vence o que a régua apurou
            "categoria": d.get("categoria", ""),
            "entra": d.get("entra", "PENDENTE"),
            "justificativa": d.get("justificativa", ""),
            "decidido_por": d.get("quem", ""), "decidido_em": d.get("quando", ""),
            # bandeiras
            "flag_tmae": "não conferido — não há extração de TMAE deste mês",
            "flag_sem_os": not campo(r, "NUMERO_OS"),
            "flag_ausente_critica": estado == "AUSENTE",
            "flag_gemea": any(g["duplicata_de_repasse"] for g in gem),
        }
        (registros if item["no_recorte"] else ampliado).append(item)

    registros.sort(key=lambda x: (x["abertura"][6:10], x["abertura"][3:5],
                                  x["abertura"][:2], x["abertura"][11:]))
    entram = [x for x in registros if x["entra"] == "SIM"]
    saida = {
        "gerado_em": data("01/01/2000") and None or "",
        "mes": a.mes, "titulo": f"{MES_NOME[mm]} de {aaaa}",
        "previa": True,
        "regra": "Indicador SEPARADO. Não soma com as 1.305 de janeiro a junho, "
                 "que estão congeladas.",
        "fontes": [os.path.basename(crit), "BASE_SS_OS_parte1.txt",
                   "BASE_SS_OS_parte2.txt"] + ([os.path.basename(a.universo)] if a.universo else []),
        "resumo": {
            "no_recorte": len(registros),
            "entram": len(entram),
            "pendentes": len([x for x in registros if x["entra"] == "PENDENTE"]),
            "fora": len([x for x in registros if x["entra"] == "NÃO"
                         or str(x["entra"]).upper().startswith("NÃO")]),
            "casaram_na_critica": len([x for x in registros if x["critica"] == "SIM"]),
            "ausentes_da_critica": len([x for x in registros if x["critica"] == "AUSENTE"]),
            "sem_os": len([x for x in registros if x["flag_sem_os"]]),
            "suspeita_auxiliar": len([x for x in registros if x["auxiliar"] != "NAO"]),
            "universo_ampliado": len(ampliado),
        },
        "registros": registros,
        "ampliado": ampliado,
        "avisos": [
            "PRÉVIA: os casos que dependem de informação de campo continuam pendentes "
            "de decisão e estão marcados como tal.",
            "Nenhum caso deste mês foi conferido contra o TMAE: não há extração de "
            "TMAE deste período. Por decisão do dono, isso não retira ninguém do indicador.",
            "Ausente da Crítica não é contraprova: a ocorrência só entra quando finaliza, "
            "e caso recente costuma aparecer na extração seguinte.",
            f"O recorte oficial do mês tem {len(registros)} SS. A régua, com filtro mais "
            f"largo, alcança mais {len(ampliado)} que citam transformador — elas estão "
            f"em 'ampliado', à vista, sem entrar na conta.",
        ],
    }
    from datetime import datetime as _dt
    saida["gerado_em"] = _dt.now().strftime("%d/%m/%Y %H:%M")
    json.dump(saida, open(a.saida, "w", encoding="utf-8"), ensure_ascii=False)
    print(f"\ngravado: {a.saida}")
    print(f"  no recorte: {len(registros)} · entram: {len(entram)} · "
          f"pendentes: {saida['resumo']['pendentes']} · fora: {saida['resumo']['fora']}")
    print(f"  universo ampliado (fora do recorte): {len(ampliado)}")


if __name__ == "__main__":
    main()
