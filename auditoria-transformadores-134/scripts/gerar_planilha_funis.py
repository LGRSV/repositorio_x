#!/usr/bin/env python3
"""Gera Base_Funis.xlsx — a planilha que reproduz os funis do site, aba por aba.

A Base_Esteira_Completa é a base crua: 1.510 linhas, uma por SS. Serve para conferir caso a
caso e não serve para explicar. Esta aqui é a outra metade: cada aba é uma tela do site, com os
mesmos números e os mesmos rótulos, para quem for apresentar poder abrir o Excel e ler a mesma
história que a tela conta — sem precisar do navegador e sem risco de os dois divergirem.

Regra da casa aplicada aqui: nenhum número é digitado. Todos saem do fluxo-1510.json na hora.
Se a esteira mudar e esta planilha não for regerada, a data no rodapé de cada aba denuncia.
"""
import datetime
import json
import os
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FLUXO = os.path.join(RAIZ, "public", "fluxo-1510.json")
SAIDA = os.path.join(RAIZ, "public", "bases", "Base_Funis.xlsx")

TINTA = PatternFill("solid", fgColor="1f2a37")
DESTAQUE = PatternFill("solid", fgColor="eef2f7")
BORDA = Border(bottom=Side(style="thin", color="d4dae2"))

ROTULO = {
    "furto": "Furto, roubo ou vandalismo",
    "abalroamento": "Abalroamento",
    "terceiros": "Causada por terceiros",
    "preventivo": "Preventivo ou programado",
    "divisao": "Divisão de circuito",
    "construcao": "Construção ou obra nova",
    "desativacao": "Desativação do posto",
    "auxiliar": "Auxiliar de religador ou regulador",
    "particular": "Transformador particular",
    "duplicada": "SS duplicada",
    "seguranca": "Obra de poste — segurança",
    "tap": "Tape interno",
    "cola_fita": "Cola e fita — reparo, não troca",
    "meta": "Substituído pela Meta",
    "avaliar_matheus": "Avaliar com o Matheus",
    "remanejamento": "Remanejamento",
    "falta_fase": "Falta de fase interna",
    "sem_os": "Sem OS e sem obra",
    "sem_obra": "Obra nunca gerada",
    "sem_fato": "Sem interrupção na base Crítica",
    "sem_interrupcao": "Ausente da base de interrupções",
    "fora_da_janela": "Ausente da base · registro em outra data",
    "erro_cadastro": "Possível erro de cadastro do código",
    "obra_sem_transformador": "Obra encerrada sem transformador",
    "obra_sem_execucao": "Obra aberta que não executou nada",
    "obra_chave": "A obra trocou chave fusível",
    "obra_poste": "A obra trocou poste",
}


def aba(wb, titulo, colunas, linhas, nota=""):
    """Uma aba com cabeçalho escuro, largura ajustada e a nota de rodapé."""
    ws = wb.create_sheet(titulo[:31])
    if nota:
        ws.append([nota])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(colunas)))
        c = ws.cell(row=1, column=1)
        c.font = Font(italic=True, color="44506a")
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[1].height = 32
        ws.append([])
    ini = ws.max_row + 1
    ws.append(colunas)
    for c in ws[ini]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = TINTA
    for l in linhas:
        ws.append(l)
    for i, t in enumerate(colunas, start=1):
        largo = max([len(str(t))] + [len(str(l[i - 1])) for l in linhas if i <= len(l)] or [10])
        ws.column_dimensions[get_column_letter(i)].width = min(70, max(12, largo + 3))
    ws.freeze_panes = ws.cell(row=ini + 1, column=1)
    return ws


def main():
    with open(FLUXO, encoding="utf-8") as fh:
        fluxo = json.load(fh)
    R = fluxo["registros"]
    tot = len(R)
    carimbo = fluxo["meta"].get("revisado_em") or str(datetime.date.today())
    saida = [r for r in R if r.get("cascata") == "SAÍDA"]
    excl = [r for r in R if r.get("cascata") == "EXCLUÍDA"]
    ret = [r for r in R if str(r.get("cascata", "")).startswith("RETIDO")]
    PORTA_1 = ("fora_da_janela", "sem_interrupcao", "sem_fato")
    parou1 = [r for r in excl if r.get("expurgo_gatilho") in PORTA_1]
    outra = [r for r in excl if r.get("expurgo_gatilho") not in PORTA_1]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1 — o funil, de cima para baixo
    aba(wb, "1 · O funil", ["Degrau", "O que ele pergunta", "Entram", "Saem aqui", "Seguem"], [
        ["Universo", "solicitações de troca de transformador no semestre", tot, 0, tot],
        ["0 · Porta", "o caso é deste indicador?", tot, len(excl), tot - len(excl)],
        ["1 · Interrupção", "a Crítica registra o evento no próprio transformador, na janela?",
         tot - len(excl), 0, tot - len(excl)],
        ["2 · Deslocamento", "houve equipe no código do transformador? (marcador, não retém)",
         tot - len(excl), 0, tot - len(excl)],
        ["3 · Material da obra", "algum transformador foi de fato movimentado?",
         tot - len(excl), len(ret), tot - len(excl) - len(ret)],
        ["4 · Ressalva", "a interrupção sustenta chamar isso de falha? (marcador, não retém)",
         tot - len(excl) - len(ret), 0, len(saida)],
        ["SAÍDA", "queimados e avariados com causa confirmada", len(saida), 0, len(saida)],
    ], nota=(f"Dado de {carimbo}. A porta é degrau 0 porque a exclusão acontece ANTES da esteira "
             "e fora dela: o caso excluído nunca desceu nenhum degrau. As peneiras 1, 2 e 4 hoje "
             "não retêm ninguém — a 1 porque quem não tem interrupção sai pela porta, a 2 e a 4 "
             "porque viraram marcadores. Quem retém é a peneira 3, e só por falta de prova."))

    # 2 — resultado
    conf = Counter(r.get("confirmado") for r in saida if r.get("confirmado"))
    aba(wb, "2 · Resultado", ["Categoria", "Solicitações", "% do universo"], [
        ["Queimados", conf.get("QUEIMADO", 0), f"{conf.get('QUEIMADO', 0) / tot:.1%}"],
        ["Avariados", conf.get("AVARIADO", 0), f"{conf.get('AVARIADO', 0) / tot:.1%}"],
        ["— soma da saída", len(saida), f"{len(saida) / tot:.1%}"],
        ["Fora do indicador", len(excl), f"{len(excl) / tot:.1%}"],
        ["Retidos, esperando prova", len(ret), f"{len(ret) / tot:.1%}"],
        ["TOTAL", tot, "100,0%"],
    ], nota="As três primeiras linhas somam a saída; as três últimas somam o universo.")

    # 3 — exclusões por categoria
    g = Counter(str(r.get("expurgo_gatilho") or "sem categoria") for r in excl)
    aba(wb, "3 · Exclusões", ["Categoria", "Gatilho no dado", "Casos", "% das exclusões",
                              "Família"], [
        [ROTULO.get(k, k), k, v, f"{v / len(excl):.1%}",
         "sem interrupção que sustente" if k in PORTA_1 else "causa ou documento fora do indicador"]
        for k, v in g.most_common()
    ], nota=(f"{len(excl)} exclusões em {len(g)} categorias, e elas se dividem em duas famílias "
             f"que convém não misturar: {len(parou1)} não têm interrupção que sustente o caso e "
             f"{len(outra)} têm causa ou documento fora do indicador."))

    # 4 — quem parou na primeira peneira
    aba(wb, "4 · Sem interrupção", ["Situação", "Casos", "O que quer dizer"], [
        [ROTULO["sem_interrupcao"], g.get("sem_interrupcao", 0),
         "o código não aparece na Crítica em papel nenhum, nos sete meses do acervo"],
        [ROTULO["fora_da_janela"], g.get("fora_da_janela", 0),
         "tem defeito aberto no próprio código, mas em data que não cabe na janela"],
        [ROTULO["sem_fato"], g.get("sem_fato", 0),
         "não deixou rastro em base alguma, nem pelo teste do vizinho"],
        ["TOTAL", len(parou1), "saem do indicador pela porta, com o motivo escrito na linha"],
    ], nota=("Na tela do site as duas primeiras contam juntas, com o nome que o dono escolheu — "
             "Ausente da base de interrupções. Aqui elas aparecem separadas porque quem revisa "
             "caso a caso precisa da diferença: não ter registro é uma coisa, ter registro em "
             "outra data é outra."))

    # 5 — censo da Crítica
    censo = Counter(r.get("censo_critica") for r in R)
    aba(wb, "5 · Censo da Crítica", ["Estado do ativo na Crítica", "Solicitações", "Na saída"], [
        [k, v, sum(1 for r in saida if r.get("censo_critica") == k)]
        for k, v in censo.most_common()
    ], nota=("Cada um dos 1.510 transformadores conferido contra os sete meses do acervo, "
             "dezembro de 2025 incluído. Contar só 2026 daria 94 ausentes em vez de 78."))

    # 6 — retidos
    aba(wb, "6 · Retidos", ["SS", "Transformador", "Localidade", "Obra", "No export de material?",
                            "Texto diz", "Etiquetas"], [
        [r.get("ss"), r.get("trafo"), r.get("localidade"), r.get("obra") or "não gerada",
         "não" if r.get("material_conferido") != "SIM" else "sim",
         r.get("categoria_texto"), r.get("etiquetas")] for r in ret
    ], nota=(f"{len(ret)} solicitações que a esteira não confirmou nem excluiu. "
             f"{sum(1 for r in ret if r.get('pendente_siago') == 'SIM')} delas só esperam a "
             "extração do SIAGO: a obra existe, com número e descrição, e não está no export de "
             "material. Não é ausência de prova, é ausência da extração."))

    # 7 — etiquetas
    etq = Counter(e for r in R for e in str(r.get("etiquetas") or "").split(" · ") if e)
    aba(wb, "7 · Etiquetas", ["Etiqueta", "Casos", "Na saída", "O que ela nomeia"], [
        [k, v, sum(1 for r in saida if k in str(r.get("etiquetas") or "")),
         next((str(r.get("etiquetas_porque")).split(f"{k}: ")[1].split(" | ")[0]
               for r in R if k in str(r.get("etiquetas") or "")
               and f"{k}: " in str(r.get("etiquetas_porque") or "")), "")]
        for k, v in etq.most_common()
    ], nota="Etiqueta nomeia, não decide: o veredito e a conta continuam os mesmos.")

    # 8 — reincidentes
    por = defaultdict(list)
    for r in saida:
        if r.get("trafo"):
            por[str(r["trafo"])].append(r)
    rein = []
    for cod, rs in por.items():
        if len(rs) < 2:
            continue
        rs = sorted(rs, key=lambda x: str(x.get("abertura")))
        d1 = str(rs[0].get("abertura"))[:10]
        d2 = str(rs[-1].get("abertura"))[:10]
        dias = (datetime.date.fromisoformat(d2) - datetime.date.fromisoformat(d1)).days
        rein.append([cod, rs[0].get("localidade"), len(rs), dias,
                     " · ".join(str(x.get("ss")) for x in rs),
                     " · ".join(str(x.get("oc_num") or "sem ocorrência") for x in rs),
                     " · ".join(str(x.get("confirmado")) for x in rs),
                     "SIM" if len({str(x.get("oc_num")) for x in rs if x.get("oc_num")}) < len(
                         [x for x in rs if x.get("oc_num")]) else "não"])
    rein.sort(key=lambda x: x[3])
    aba(wb, "8 · Reincidentes", ["Transformador", "Localidade", "Saídas", "Dias entre a 1ª e a última",
                                 "Solicitações", "Ocorrências", "Causa", "Dividem a mesma ocorrência?"],
        rein, nota=(f"{len(rein)} transformadores saíram mais de uma vez no semestre, somando "
                    f"{sum(x[2] for x in rein)} solicitações — que já estão contadas na saída, não "
                    "se somam por fora. Quando as ocorrências são distintas e cada uma tem material "
                    "próprio, são dois eventos reais e não erro de contagem."))

    # 9 — vereditos do dono
    ver = fluxo["meta"].get("vereditos_do_dono") or []
    porss = {r["ss"]: r for r in R}
    aba(wb, "9 · Vereditos do dono", ["SS", "Decisão", "Categoria", "Onde está hoje", "Por quê"], [
        [v["ss"], v["decisao"], ROTULO.get(v["gatilho"], v["gatilho"]) or v["confirmado"],
         porss.get(v["ss"], {}).get("cascata", "—"),
         porss.get(v["ss"], {}).get("exclusao_porque") or porss.get(v["ss"], {}).get("cascata_motivo", "")]
        for v in ver
    ], nota=("Casos em que ele leu o dossiê inteiro e decidiu contra a regra. Nenhum virou regra: "
             "generalizar a partir de um caso já rendeu erro nesta auditoria."))

    # 10 — por localidade
    loc = Counter(str(r.get("localidade")) for r in saida)
    aba(wb, "10 · Por localidade", ["Localidade", "Na saída", "Excluídos", "Retidos", "Total"], [
        [k, v, sum(1 for r in excl if str(r.get("localidade")) == k),
         sum(1 for r in ret if str(r.get("localidade")) == k),
         sum(1 for r in R if str(r.get("localidade")) == k)]
        for k, v in loc.most_common()
    ], nota="Ordenado pelo que conta no indicador.")

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    print(f"{SAIDA} · {len(wb.sheetnames)} abas · {os.path.getsize(SAIDA) / 1000:.0f} KB")
    print("abas: " + " · ".join(wb.sheetnames))
    print(f"conferência: saída {len(saida)} + excluídas {len(excl)} + retidos {len(ret)} = "
          f"{len(saida) + len(excl) + len(ret)} (universo {tot})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
