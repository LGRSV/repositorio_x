#!/usr/bin/env python3
"""Junta o que cada base diz sobre FURTO, ABALROAMENTO e DANO DE TERCEIRO e grava terceiros.json.

A pergunta dele: "no universo das 1.305, a Crítica traz ou o campo traz algum como abalroado,
dano causado por terceiro ou furtado?" — e depois, quando respondi olhando pouca coisa: "além
das descrições da obra tem mais algum outro campo que confirma isso?".

Tem, e a resposta só apareceu varrendo TODAS as colunas de TODAS as bases. São seis campos, em
quatro bases diferentes, e cada um fala de uma coisa:

    OS · PROVÁVEL_MOTIVO_DO_DEFEITO   o formulário que a EQUIPE preenche no poste. É a voz de
                                      quem viu o equipamento — e acusa casos que nenhuma outra
                                      base menciona, inclusive um VANDALISMO
    OS · DEFEITO e DEFEITO_SS         a classificação do defeito na SS e na OS
    OS · ORIGEM                       a origem declarada da solicitação
    obra · DESCRICAO                  o carimbo contábil: "FURTO DE BENS", "DANO(S) CAUSADO POR
                                      TERCEIROS" — preenchido DEPOIS da execução
    obra · DESCRICAO_OBRA             o campo livre do cadastro, às vezes com o código do ativo
    Crítica e TMAE · causa/subcausa   ABALROADO, CAUSADA POR TERCEIROS/INCÊNDIO

POR QUE CONTAR AS FONTES. Uma fonte sozinha não condena — o formulário pode ter sido preenchido
no chute, e o carimbo da obra é escolha contábil. Quando quatro ou cinco campos independentes
dizem a mesma coisa, é outra conversa. Por isso cada caso sai daqui com a lista de quem
confirma e o número, e a tela ordena por ele.

NADA AQUI MOVE NINGUÉM. É leitura: os 1.305 não se mexem.

Uso:  python3 scripts/gerar_terceiros.py
"""
import collections
import datetime
import json
import os
import re

import openpyxl

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FLUXO = os.path.join(RAIZ, "public", "fluxo-1510.json")
OS_XLSX = os.path.join(RAIZ, "public", "bases", "originais", "Original_OS.xlsx")
OBRAS = os.path.join(RAIZ, "public", "bases", "originais", "Original_Cadastro_Obras.xlsx")
SAIDA = os.path.join(RAIZ, "public", "terceiros.json")

PAT = re.compile(r"ABALROA|FURT|ROUB|TERCEIRO|VANDAL|DEPREDA|INCENDIO|INC[ÊE]NDIO", re.I)

ROTULO = {
    "PROVÁVEL_MOTIVO_DO_DEFEITO": "OS · provável motivo do defeito (formulário da equipe)",
    "DEFEITO": "OS · defeito",
    "DEFEITO_SS": "SS · defeito",
    "ORIGEM": "SS · origem",
    "DESCRICAO": "Obra · descrição (projeto contábil)",
    "DESCRICAO_OBRA": "Obra · descrição livre do cadastro",
    "oc_causa": "Crítica · causa",
    "oc_sub": "Crítica · subcausa",
    "at_causa": "TMAE · causa",
    "at_sub": "TMAE · subcausa",
}


def t(v):
    return "" if v is None else str(v).strip()


def main():
    with open(FLUXO, encoding="utf-8") as fh:
        fluxo = json.load(fh)
    regs = fluxo["registros"]
    por_os = {t(r.get("os")): t(r.get("ss")) for r in regs if t(r.get("os"))}
    por_obra = {t(r.get("obra")).lstrip("0"): t(r.get("ss")) for r in regs if t(r.get("obra"))}
    fontes = collections.defaultdict(dict)

    # ---------- a base da OS: o formulário da equipe e os campos de defeito
    wb = openpyxl.load_workbook(OS_XLSX, read_only=True, data_only=True)
    ws = wb["BASE SS_OS"]
    it = ws.iter_rows(values_only=True)
    cab = list(next(it))
    ic = {c: i for i, c in enumerate(cab) if c}
    for row in it:
        ss = por_os.get(t(row[ic["NUMERO_OS"]]) if ic["NUMERO_OS"] < len(row) else "")
        if not ss:
            continue
        for c in ("PROVÁVEL_MOTIVO_DO_DEFEITO", "DEFEITO", "DEFEITO_SS", "ORIGEM"):
            v = t(row[ic[c]]) if c in ic and ic[c] < len(row) else ""
            if v and PAT.search(v):
                fontes[ss][ROTULO[c]] = v
    wb.close()

    # ---------- o cadastro de obras: o carimbo contábil e o campo livre
    wb2 = openpyxl.load_workbook(OBRAS, read_only=True, data_only=True)
    ws2 = wb2["Export"]
    it2 = ws2.iter_rows(values_only=True)
    cab2 = list(next(it2))
    ic2 = {c: i for i, c in enumerate(cab2) if c}
    for row in it2:
        ss = por_obra.get(t(row[ic2["NUM_OBRA"]]).lstrip("0"))
        if not ss:
            continue
        for c in ("DESCRICAO", "DESCRICAO_OBRA"):
            v = t(row[ic2[c]]) if c in ic2 and ic2[c] < len(row) else ""
            if v and PAT.search(v):
                fontes[ss][ROTULO[c]] = v[:160]
    wb2.close()

    # ---------- a Crítica e o TMAE, que já viajam no fluxo
    for r in regs:
        ss = t(r.get("ss"))
        for c in ("oc_causa", "oc_sub", "at_causa", "at_sub"):
            v = t(r.get(c))
            if v and PAT.search(v):
                fontes[ss][ROTULO[c]] = v

    saida = {ss: {"n": len(d), "fontes": [{"campo": k, "valor": v} for k, v in sorted(d.items())]}
             for ss, d in fontes.items()}
    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    with open(SAIDA, "w", encoding="utf-8") as fh:
        json.dump({"gerado_em": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
                   "por_ss": saida}, fh, ensure_ascii=False)

    c = collections.Counter(v["n"] for v in saida.values())
    print(SAIDA)
    print(f"  solicitações com furto, abalroamento ou dano de terceiro em alguma base: {len(saida)}")
    print(f"  por número de fontes que confirmam: {dict(sorted(c.items()))}")
    for ss, d in sorted(saida.items(), key=lambda x: -x[1]["n"])[:6]:
        print(f"   {d['n']} fontes · {ss}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
