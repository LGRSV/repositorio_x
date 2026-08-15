"""As obras abertas em volta do transformador, pela coordenada.

Ele perguntou: "em obras teria como colocar as obras abertas naquela região, tipo pela
coordenada, histórico de obras abertas na região?" — tem.

A extração de obras traz 124.084 linhas e todas com coordenada. Duas coisas precisavam
ser descobertas antes de confiar nelas, e as duas foram medidas, não supostas:

  1. AS COLUNAS ESTÃO TROCADAS. "COORDENADA X" guarda o NORTE (8,9 milhões) e
     "COORDENADA Y" guarda o LESTE (~800 mil). Quem ler pelo nome inverte o mapa.
  2. A PROJEÇÃO É UTM 22S (EPSG:31982), inclusive para a faixa leste do estado, onde a
     regra do fuso mandaria 23S. Descobri comparando 1.447 obras que a auditoria já liga a
     um ativo com coordenada conhecida: o erro mediano em 22S deu ZERO metro, e em 23S deu
     657 km. Não é estimativa — é a mesma fonte dos dois lados.

O resto é vizinhança: para cada transformador da auditoria, as obras dentro de 2 km, da
mais perto para a mais longe, com o que a obra é e em que estado ela está. Serve para a
pergunta que não se responde caso a caso — "estavam mexendo na rede aqui em volta?".
"""

import json
import math
import os
from collections import defaultdict

from openpyxl import load_workbook
from pyproj import Transformer

APP = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "public")
OBRAS = "/root/.claude/uploads/74dc9c64-5026-54ee-a81e-173d2f38a735/4368efa3-data__20260807T090818.918.xlsx"
RAIO = 2000          # metros
MAX_POR_ATIVO = 40   # as mais próximas; além disso vira ruído
CELULA = 2000        # a grade que evita comparar 1.894 × 124.084


def txt(v):
    return "" if v is None else str(v).strip()


# ── onde está cada transformador, em metros ────────────────────────────────
para_utm = Transformer.from_crs("EPSG:4326", "EPSG:31982", always_xy=True)
ativos = {}
fluxo = json.load(open(f"{APP}/fluxo-1510.json", encoding="utf-8"))
fontes = [fluxo["registros"]]
for arq in ("julho-2026.json", "agosto-2026.json"):
    d = json.load(open(f"{APP}/{arq}", encoding="utf-8"))
    fontes += [d["registros"], d.get("ampliado", [])]
for grupo in fontes:
    for r in grupo:
        t = txt(r.get("trafo"))
        if not t or t in ativos:
            continue
        lat, lon = r.get("lat"), r.get("lon")
        if lat and lon:
            e, n = para_utm.transform(float(lon), float(lat))
            ativos[t] = (e, n)
print(f"{len(ativos)} transformadores com coordenada conhecida")

grade = defaultdict(list)
for t, (e, n) in ativos.items():
    grade[(int(e // CELULA), int(n // CELULA))].append(t)

# ── as obras, uma passada só ───────────────────────────────────────────────
wb = load_workbook(OBRAS, read_only=True)
ws = wb["Export"]
it = ws.iter_rows(values_only=True)
cab = [txt(c) for c in next(it)]
J = {c: i for i, c in enumerate(cab)}


def g(linha, coluna):
    i = J.get(coluna, -1)
    return txt(linha[i]) if 0 <= i < len(linha) else ""


perto = defaultdict(list)
lidas = com_coord = 0
for linha in it:
    if linha is None or all(v is None for v in linha):
        continue
    lidas += 1
    try:
        # o NORTE está na coluna chamada X e o LESTE na chamada Y — ver o cabeçalho
        norte, leste = float(g(linha, "COORDENADA X")), float(g(linha, "COORDENADA Y"))
    except ValueError:
        continue
    if not (7_000_000 < norte < 10_500_000 and 100_000 < leste < 1_000_000):
        continue          # zero, 111111/3333333 e outros lixos de digitação
    com_coord += 1
    cx, cy = int(leste // CELULA), int(norte // CELULA)
    candidatos = []
    for dx in (-1, 0, 1):
        for dy in (-1, 0, 1):
            candidatos += grade.get((cx + dx, cy + dy), [])
    if not candidatos:
        continue
    obra = {
        "o": g(linha, "NUM_OBRA"), "s": g(linha, "DSC_STATUS"),
        "oc": g(linha, "DSC_OCORRENCIA"), "aic": g(linha, "AIC"),
        "c": g(linha, "CLASS_OBRA"), "t": g(linha, "CONST_MANUT"),
        "d": g(linha, "DESCRICAO") or g(linha, "DESCRICAO_OBRA"),
        "e": g(linha, "EMPREITEIRA_PROJETO"), "a": g(linha, "AREA"),
        "p": g(linha, "DATA_PARALISACAO"), "mp": g(linha, "MOTV_PARALIZACAO"),
        "os": g(linha, "NUM_OS"), "r": g(linha, "SGL_RESPONSAVEL"),
    }
    for t in candidatos:
        e, n = ativos[t]
        dist = math.hypot(leste - e, norte - n)
        if dist <= RAIO:
            perto[t].append({**obra, "m": round(dist)})

for t in perto:
    perto[t].sort(key=lambda x: x["m"])
    del perto[t][MAX_POR_ATIVO:]

n_obras = sum(len(v) for v in perto.values())
print(f"{lidas} obras lidas · {com_coord} com coordenada plausível")
print(f"{len(perto)} transformadores têm obra dentro de {RAIO} m · {n_obras} ligações")

# ── entra no mesmo arquivo do histórico, que já é pedido sob demanda ───────
cam = f"{APP}/historico-ativo.json"
h = json.load(open(cam, encoding="utf-8"))
for t, v in h["por_ativo"].items():
    v["obras"] = perto.get(t, [])
h["obras_raio_m"] = RAIO
h["obras_lidas"] = lidas
json.dump(h, open(cam, "w"), ensure_ascii=False, separators=(",", ":"))
print(f"gravado: {cam} ({os.path.getsize(cam)/1e6:.1f} MB)")

campeao = max(perto.items(), key=lambda kv: len(kv[1]))
print(f"\nativo com mais obras em volta: {campeao[0]} — {len(campeao[1])}")
for o in campeao[1][:5]:
    print(f"   {o['m']:>5} m  obra {o['o']}  {o['s'][:34]:34s} {o['c'][:18]}")
